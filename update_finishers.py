#!/usr/bin/env python3
"""
Utility to update finisher counts in the Excel file and regenerate dashboard.

Usage:
    python update_finishers.py "Race Name" DISTANCE YEAR COUNT
    python update_finishers.py "TCS London Marathon" MARATHON 2012 36000

DISTANCE must be one of: MARATHON, SEMI, 10KM, AUTRE
"""
import sys
import openpyxl
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / "Suivi_Finishers_Monde_10k_-_21k_-_42k_HISTORIQUE.xlsx"


def update(race_name, distance, year, count):
    year = int(year)
    count = int(count)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["ALL"]

    # Find year column
    header = [cell.value for cell in ws[1]]
    year_col = None
    for i, h in enumerate(header):
        if isinstance(h, (int, float)) and int(h) == year:
            year_col = i + 1  # 1-indexed
            break

    if year_col is None:
        print(f"ERREUR: Annee {year} non trouvee dans les colonnes du fichier.")
        print(f"Colonnes disponibles: {[int(h) for h in header if isinstance(h, (int, float))]}")
        return False

    # Find race row
    found = False
    for row in ws.iter_rows(min_row=2):
        cell_dist = str(row[2].value).strip() if row[2].value else ""
        cell_race = str(row[3].value).strip() if row[3].value else ""
        if cell_dist == distance and cell_race == race_name:
            old_val = row[year_col - 1].value
            # NEVER overwrite existing data
            if old_val is not None and str(old_val).strip() != '':
                print(f"[SKIP] {race_name} ({distance}) {year}: deja rempli ({old_val}), pas de modification.")
                return True
            row[year_col - 1].value = count
            found = True
            print(f"[NOUVEAU] {race_name} ({distance}) {year}: {old_val} -> {count}")
            break

    if not found:
        print(f"ERREUR: Course '{race_name}' ({distance}) non trouvee dans l'onglet ALL.")
        print("Courses disponibles pour cette distance:")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[2]).strip() == distance:
                print(f"  - {row[3]}")
        return False

    wb.save(EXCEL_FILE)
    print(f"Excel sauvegarde.")
    return True


if __name__ == "__main__":
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    race = sys.argv[1]
    dist = sys.argv[2]
    year = sys.argv[3]
    count = sys.argv[4]

    if dist not in ("MARATHON", "SEMI", "10KM", "AUTRE"):
        print(f"ERREUR: Distance '{dist}' invalide. Utiliser MARATHON, SEMI ou 10KM.")
        sys.exit(1)

    success = update(race, dist, int(year), int(count))
    if not success:
        sys.exit(1)
