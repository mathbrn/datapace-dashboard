#!/usr/bin/env python3
"""
scraper_archives.py — Scraper Playwright dedie pour decouverte dynamique
des IDs historiques par event.

Pour chaque event dans event_platform_map.json avec une plateforme connue,
tente de recuperer les finishers pour TOUTES les annees 2019-2024 :

1. ChronoRace modern : visite la page d'event, collecte les ctx historiques
   via Playwright (intercepte /api/results/table/search/{ctx}/LIVE1)
2. NYRR : boucle sur codes {YY}{CODE} pour 2019-2024
3. Sporthive : search API par nom d'event -> liste de sporthive_ids -> count
4. RaceResult : pas de multi-annee via ID unique -> skip

USAGE:
    python scraper_archives.py --events N     # traiter les N events les plus grands
    python scraper_archives.py --dry-run      # ne pas ecrire dans l'Excel
"""
import argparse, json, sys, re, time
from datetime import date, datetime
from pathlib import Path
import requests
import openpyxl
from playwright.sync_api import sync_playwright

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

SCRIPT_DIR = Path(__file__).parent
XLSX = SCRIPT_DIR / "Suivi_Finishers_Monde_10k_-_21k_-_42k_HISTORIQUE.xlsx"
EPM = SCRIPT_DIR / "event_platform_map.json"
LOG_DIR = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

H = {"User-Agent":"Mozilla/5.0","Accept":"application/json"}
NYRR_H = {**H,"Content-Type":"application/json","Origin":"https://results.nyrr.org","Referer":"https://results.nyrr.org/"}

YEAR_COL_START = 5

def is_round(v):
    return isinstance(v, int) and v >= 500 and v % 500 == 0

def load_epm():
    with open(EPM, encoding="utf-8") as f:
        return json.load(f)

def load_xlsx():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["ALL"]
    row_map = {}
    for i in range(2, ws.max_row+1):
        race = ws.cell(i,4).value
        if race:
            row_map.setdefault(race, []).append({"row":i,"dist":str(ws.cell(i,3).value or "")})
    return wb, ws, row_map

def scrape_nyrr(event_name, code_prefix):
    """NYRR: try codes {YY}{PREFIX} for 2019-2024."""
    results = {}
    for yr in range(19, 25):
        code = f"{yr}{code_prefix}"
        try:
            r = requests.post("https://rmsprodapi.nyrr.org/api/v2/runners/finishers-filter",
                              json={"eventCode":code,"pageNum":1,"pageSize":1,"sortColumn":"overallPlace","sortDescending":False},
                              headers=NYRR_H, timeout=10)
            if r.status_code == 200:
                d = r.json()
                cnt = d.get("totalItems",0)
                if cnt > 100 and not is_round(cnt):
                    items = d.get("items",[])
                    winner = f"{items[0].get('firstName','')} {items[0].get('lastName','')} {items[0].get('overallTime','')}" if items else ""
                    results[2000+yr] = {"count":cnt,"winner":winner,"source":f"nyrr/{code}"}
        except: pass
    return results

def scrape_chronorace_archive(event_name, event_id_modern=None):
    """Playwright: open archives page, intercept /api/results/table/search/ calls, collect ctx."""
    ctxs_found = []
    # URL pattern for modern chronorace event
    if not event_id_modern:
        return {}
    url = f"https://prod.chronorace.be/Classements/Classement.aspx?eventId={event_id_modern}"
    with sync_playwright() as p:
        b = p.chromium.launch(headless=True)
        page = b.new_page()
        def on_req(req):
            if "/api/results/table/search/" in req.url:
                ctxs_found.append(req.url)
        page.on("request", on_req)
        try:
            page.goto(url, wait_until="networkidle", timeout=20000)
            page.wait_for_timeout(2000)
        except: pass
        b.close()
    # Extract ctx from URLs and query each for Count
    results = {}
    for u in ctxs_found:
        m = re.search(r"/table/search/(\d{8}_[a-z0-9_]+)/(\w+)", u)
        if not m: continue
        ctx, rep = m.group(1), m.group(2)
        year = int(ctx[:4])
        try:
            r = requests.get(f"https://results.chronorace.be/api/results/table/search/{ctx}/{rep}?srch=&pageSize=1&fromRecord=0",
                             headers=H, timeout=8)
            if r.status_code == 200:
                d = r.json()
                cnt = d.get("Count",0)
                if cnt > 100 and not is_round(cnt):
                    results.setdefault(year, {"counts":{},"winners":{}})
                    results[year]["counts"][rep] = cnt
                    # Extract winner
                    grps = d.get("Groups",[{}])
                    srows = grps[0].get("SlaveRows",[])
                    if srows:
                        first = srows[0]
                        nm = re.search(r"<b>([^<]+)</b>", str(first[2])) if len(first)>2 else None
                        tm = re.search(r"<b>([^<]+)</b>", str(first[5])) if len(first)>5 else None
                        if nm and tm:
                            results[year]["winners"][rep] = f"{nm.group(1).strip()} {tm.group(1).strip()}"
        except: pass
    return results

def scrape_sporthive_search(event_name):
    """Search Sporthive for all editions matching name."""
    try:
        r = requests.get(f"https://eventresults-api.speedhive.com/sporthive/events?searchQuery={event_name}",
                         headers={"User-Agent":"Mozilla/5.0","Accept":"application/json"}, timeout=15)
        if r.status_code == 200:
            data = r.json()
            results = {}
            events = data if isinstance(data, list) else (data.get("events") or data.get("items") or [])
            for e in events[:30]:
                eid = e.get("id") or e.get("eventId")
                date_s = str(e.get("date","") or e.get("startDate",""))[:4]
                if not eid or not date_s.isdigit(): continue
                year = int(date_s)
                # Get races for this event
                rr = requests.get(f"https://eventresults-api.speedhive.com/sporthive/events/{eid}/races",
                                  headers={"User-Agent":"Mozilla/5.0"}, timeout=10)
                if rr.status_code == 200:
                    races = rr.json() if isinstance(rr.json(), list) else []
                    total = sum(race.get("classificationsCount",0) for race in races if isinstance(race, dict))
                    if total > 100 and not is_round(total):
                        results.setdefault(year, []).append({"eid":eid,"count":total})
            return results
    except: pass
    return {}

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--events", type=int, default=10, help="Top N events to process")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    epm = load_epm()
    wb, ws, row_map = load_xlsx()

    # Identify events with mapped platform
    candidates = []
    for key, meta in epm.items():
        name = meta.get("name")
        if not name: continue
        platform = meta.get("platform")
        # Compute peak finishers across all rows of this event
        rows = row_map.get(name, [])
        peak = 0
        empty_recent = []
        for rinfo in rows:
            r = rinfo["row"]
            for c in range(YEAR_COL_START, 32):
                y = 2000+(c-YEAR_COL_START)
                v = ws.cell(r, c).value
                if isinstance(v,(int,float)) and v > peak: peak = int(v)
                if v in (None,"") and 2019 <= y <= 2024:
                    empty_recent.append((y, rinfo["dist"]))
        if empty_recent and peak > 0:
            candidates.append({"name":name,"platform":platform,"meta":meta,"peak":peak,"empty":empty_recent})

    # Only keep events with supported platforms
    SUPPORTED = {"chronorace","nyrr","sporthive"}
    candidates = [c for c in candidates if c["platform"] in SUPPORTED]
    candidates.sort(key=lambda x: -x["peak"])
    to_process = candidates[:args.events]
    print(f"Processing top {len(to_process)} events (supported platforms only)")

    applied = []
    all_results = {}
    for c in to_process:
        name = c["name"]
        platform = c["platform"]
        meta = c["meta"]
        print(f"\n=== {name} ({platform}) peak={c['peak']} empty={len(c['empty'])} ===")
        res = {}
        if platform == "nyrr":
            prefix = meta.get("nyrr_code_prefix","")
            if prefix:
                res = scrape_nyrr(name, prefix)
        elif platform in ("chronorace",):
            eid = meta.get("chronorace_event_id")
            if eid:
                res = scrape_chronorace_archive(name, eid)
        elif platform == "sporthive":
            res = scrape_sporthive_search(name)
        else:
            print(f"  [SKIP] platform {platform} not supported by scraper yet")
            continue

        all_results[name] = res
        # Try to apply to xlsx
        if res:
            for year, data in res.items():
                # Format varies per platform — handle all
                cnt = None
                if isinstance(data, dict):
                    if "count" in data: cnt = data["count"]
                    elif "counts" in data:
                        counts = data["counts"]
                        # If LIVE1 + LIVE2 exist → prefer sum for AUTRE
                        if "LIVE1" in counts:
                            cnt = counts["LIVE1"] + counts.get("LIVE2", 0)
                elif isinstance(data, list) and data:
                    cnt = data[0].get("count")
                if not cnt or is_round(cnt): continue
                # Find matching row for this event — pick first distance available (ambiguous but pragmatic)
                rows = row_map.get(name, [])
                if not rows: continue
                # Pick row with matching AUTRE first, else first
                target_row = None
                for rinfo in rows:
                    if rinfo["dist"] == "AUTRE": target_row = rinfo; break
                if not target_row: target_row = rows[0]
                col = YEAR_COL_START + (year - 2000)
                cur = ws.cell(target_row["row"], col).value
                if cur in (None, ""):
                    if not args.dry_run:
                        ws.cell(target_row["row"], col).value = cnt
                    applied.append({"event":name,"dist":target_row["dist"],"year":year,"value":cnt,"platform":platform})
                    print(f"  [APPLY] {name} {target_row['dist']} {year} = {cnt}")

    if not args.dry_run and applied:
        wb.save(XLSX)
    print(f"\nTotal applied: {len(applied)}")

    with open(LOG_DIR / f"scraper_archives_{date.today()}.json","w",encoding="utf-8") as f:
        json.dump({"applied":applied,"all_results":all_results}, f, indent=2, ensure_ascii=False)

if __name__ == "__main__":
    main()
