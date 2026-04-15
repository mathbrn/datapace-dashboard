#!/usr/bin/env python3
"""
Import Benchmark_Courses_Monde.xlsx → Suivi_Finishers + dashboard.

Étape 1 : exploration & rapport (--analyze)
Étape 2 : web search dates manquantes (--fetch-dates)
Étape 3 : dry-run complet (--dry-run)
Étape 4 : import réel (--import)
"""
import argparse, json, sys, unicodedata
from pathlib import Path
from difflib import SequenceMatcher
import openpyxl

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

SCRIPT_DIR = Path(__file__).parent
SRC = SCRIPT_DIR / "Benchmark_Courses_Monde.xlsx"
DST = SCRIPT_DIR / "Suivi_Finishers_Monde_10k_-_21k_-_42k_HISTORIQUE.xlsx"

REGION_MAP = {
    "Europe": "Europe", "Asie": "Asie",
    "Amérique du Nord": "Amérique du Nord", "Amerique du Nord": "Amérique du Nord",
    "Amérique du Sud": "Amérique du Sud", "Amerique du Sud": "Amérique du Sud",
    "Océanie": "Océanie", "Oceanie": "Océanie",
    "Moyen-Orient": "Moyen-Orient", "Afrique": "Afrique",
    "Zagreb": "Europe",  # anomaly
}

MOIS = {1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin",
        7: "Juillet", 8: "Août", 9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"}


def norm(s):
    return unicodedata.normalize("NFKD", (s or "").lower()).encode("ascii", "ignore").decode().strip()


def load_existing_event_names():
    wb = openpyxl.load_workbook(DST, read_only=True)
    ws = wb["ALL"]
    names = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = list(row); continue
        d = dict(zip(headers, row))
        race = str(d.get("Race", "") or "").strip()
        if race:
            names.append(race)
    wb.close()
    return names


def parse_count(v):
    """Parse a cell value as a count, return None if not numeric."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace(" ", "")
        if s in ("", "X", "x"):
            return None
        # Skip formulas
        if s.startswith("="):
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return int(v)
    return None


def is_distance_present(v):
    """True if column J/K/L/M has a value (X, x, or numeric)."""
    if v is None:
        return False
    if isinstance(v, str):
        s = v.strip()
        return s != ""
    return True


def explore():
    """Étape 1 — Rapport complet."""
    existing = load_existing_event_names()
    existing_norm = [norm(n) for n in existing]
    print(f"Dashboard actuel: {len(existing)} courses")

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["ALL"]

    total_rows = 0
    trails = 0
    duplicates = 0
    new_events = 0
    missing_dates = 0
    by_region = {}
    by_distance = {"MARATHON": 0, "SEMI": 0, "10KM": 0, "5KM": 0, "AUTRE": 0}
    finishers_2025 = 0
    new_event_list = []

    for r in range(3, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        if not row or not row[4]:  # E = Événement
            continue
        total_rows += 1
        pays = row[0]
        region = row[1]
        ville = row[3]
        event = str(row[4]).strip()
        date = row[5]
        type_course = str(row[8] or "").strip().lower()  # I = Type

        # Skip Trail
        if "trail" in type_course:
            trails += 1
            continue

        # Skip already in dashboard
        ev_norm = norm(event)
        is_dup = False
        for ex_norm in existing_norm:
            if ex_norm == ev_norm:
                is_dup = True
                break
            ratio = SequenceMatcher(None, ev_norm, ex_norm).ratio()
            if ratio > 0.82:
                is_dup = True
                break
        if is_dup:
            duplicates += 1
            continue

        new_events += 1
        if not date or not hasattr(date, "month"):
            missing_dates += 1
        region_clean = REGION_MAP.get(str(region or "").strip(), region)
        by_region[region_clean] = by_region.get(region_clean, 0) + 1

        # Determine distances present
        distances_for_event = []
        if is_distance_present(row[9]):  # J = Marathon
            distances_for_event.append(("MARATHON", parse_count(row[9])))
            by_distance["MARATHON"] += 1
        if is_distance_present(row[10]):  # K = Semi
            distances_for_event.append(("SEMI", parse_count(row[10])))
            by_distance["SEMI"] += 1
        if is_distance_present(row[11]):  # L = 10K
            distances_for_event.append(("10KM", parse_count(row[11])))
            by_distance["10KM"] += 1
        if is_distance_present(row[12]):  # M = 5K
            distances_for_event.append(("5KM", parse_count(row[12])))
            by_distance["5KM"] += 1
        if is_distance_present(row[13]):  # N = Autres
            v = str(row[13] or "").strip()
            if v.lower() != "x":
                distances_for_event.append(("AUTRE", parse_count(row[13])))
                by_distance["AUTRE"] += 1

        # 2025 finishers (col P = index 15)
        f2025 = parse_count(row[15])
        has_f2025 = f2025 is not None and f2025 % 500 != 0 and f2025 > 100
        if has_f2025:
            finishers_2025 += 1

        date_month = date.month if hasattr(date, "month") else None
        new_event_list.append({
            "pays": pays, "region": region_clean, "ville": ville,
            "name": event, "date": date_month,
            "year_creation": row[19],
            "distances": distances_for_event,
            "finishers_2025": f2025 if has_f2025 else None,
        })

    wb.close()

    print()
    print("=" * 70)
    print("RAPPORT D'ANALYSE — Benchmark_Courses_Monde.xlsx")
    print("=" * 70)
    print(f"Lignes source totales:           {total_rows}")
    print(f"Trails (skipped):                {trails}")
    print(f"Doublons existants (skipped):    {duplicates}")
    print(f"Nouveaux événements à importer:  {new_events}")
    print(f"Sans date (web search requis):   {missing_dates}")
    print(f"Avec finishers 2025 valides:     {finishers_2025}")
    print()
    print("Par région:")
    for region, n in sorted(by_region.items(), key=lambda x: -x[1]):
        print(f"  {region:25s}: {n}")
    print()
    print("Par distance (lignes dashboard à créer):")
    total_lines = 0
    for d, n in by_distance.items():
        print(f"  {d:10s}: {n}")
        total_lines += n
    print(f"  TOTAL LIGNES À CRÉER: {total_lines}")
    print()
    # Save the candidates
    out = SCRIPT_DIR / "_benchmark_candidates.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump([{
            "pays": e["pays"], "region": e["region"], "ville": e["ville"],
            "name": e["name"],
            "date_month": e["date"],
            "year_creation": e["year_creation"],
            "distances": [(d, c) for d, c in e["distances"]],
            "finishers_2025": e["finishers_2025"],
        } for e in new_event_list], f, indent=2, ensure_ascii=False, default=str)
    print(f"Candidats sauvegardés: {out}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--analyze", action="store_true")
    parser.add_argument("--fetch-dates", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--import", dest="do_import", action="store_true")
    args = parser.parse_args()
    if args.analyze or not (args.fetch_dates or args.dry_run or args.do_import):
        explore()
