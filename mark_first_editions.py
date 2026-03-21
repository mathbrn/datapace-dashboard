#!/usr/bin/env python3
"""Mark cells before each event's first edition with 'x' in Excel."""
import openpyxl

FIRST_EDITIONS = {
    # MARATHONS
    ("Tokyo Marathon", "MARATHON"): 2007,
    ("TCS London Marathon", "MARATHON"): 1981,
    ("Schneider Electric Marathon de Paris", "MARATHON"): 1976,
    ("Boston Marathon", "MARATHON"): 1897,
    ("Adidas Manchester Marathon", "MARATHON"): 2012,
    ("NN Marathon Rotterdam", "MARATHON"): 1981,
    ("Vienna City Marathon", "MARATHON"): 1984,
    ("Sanlam Cape Town Marathon", "MARATHON"): 2014,
    ("Copenhagen Marathon", "MARATHON"): 1980,
    ("Prague International Marathon", "MARATHON"): 1995,
    ("BMW Berlin Marathon", "MARATHON"): 1974,
    ("Bank of America Chicago Marathon", "MARATHON"): 1977,
    ("Marine Corps Marathon", "MARATHON"): 1976,
    ("TCS Amsterdam Marathon", "MARATHON"): 1975,
    ("TCS New York City Marathon", "MARATHON"): 1970,
    ("Valencia Marathon Trinidad Alfonso Zurich", "MARATHON"): 1981,
    ("Honolulu Marathon", "MARATHON"): 1973,
    ("Flying Pig Marathon", "MARATHON"): 1999,
    ("San Francisco Marathon", "MARATHON"): 1977,
    ("Austin Marathon", "MARATHON"): 1992,
    ("Miami Marathon", "MARATHON"): 2003,
    ("Detroit Marathon", "MARATHON"): 1978,
    ("Cowtown Marathon", "MARATHON"): 1978,
    ("Shamrock Marathon", "MARATHON"): 1973,
    ("Indianapolis Monumental Marathon", "MARATHON"): 2008,
    ("Columbus Marathon", "MARATHON"): 1980,
    ("Baltimore Running Festival", "MARATHON"): 2001,
    ("Eugene Marathon", "MARATHON"): 2007,
    ("California International Marathon", "MARATHON"): 1983,
    ("Fargo Marathon", "MARATHON"): 2005,
    ("Portland Marathon", "MARATHON"): 1972,
    ("Haspa Marathon Hamburg", "MARATHON"): 1986,
    ("Mainova Frankfurt Marathon", "MARATHON"): 1981,
    ("Stockholm Marathon", "MARATHON"): 1979,
    ("Milano Marathon", "MARATHON"): 2000,
    ("Brighton Marathon", "MARATHON"): 2010,
    ("Shanghai Marathon", "MARATHON"): 1996,
    ("Beijing Marathon", "MARATHON"): 1981,
    ("Osaka Marathon", "MARATHON"): 2011,
    ("Gold Coast Marathon", "MARATHON"): 1979,
    ("Istanbul Marathon", "MARATHON"): 1979,
    ("Athens Marathon", "MARATHON"): 1972,
    ("Seoul Marathon", "MARATHON"): 1931,
    ("ASICS Los Angeles Marathon", "MARATHON"): 1986,
    ("Chevron Houston Marathon", "MARATHON"): 1972,
    ("Medtronic Twin Cities Marathon", "MARATHON"): 1982,
    ("Richmond Marathon", "MARATHON"): 1978,
    ("BMW Dallas Marathon", "MARATHON"): 1971,
    ("Long Beach Marathon", "MARATHON"): 1982,
    ("St. Jude Memphis Marathon", "MARATHON"): 2000,
    ("Oklahoma City Memorial Marathon", "MARATHON"): 2001,
    ("San Antonio Marathon", "MARATHON"): 2002,
    ("Denver Colfax Marathon", "MARATHON"): 2006,
    ("Standard Chartered Singapore Marathon", "MARATHON"): 2002,
    ("Taipei Marathon", "MARATHON"): 1986,
    ("Tata Mumbai Marathon", "MARATHON"): 2004,
    ("Rock 'n' Roll Arizona Marathon", "MARATHON"): 2004,
    ("Hoag OC Marathon", "MARATHON"): 2005,
    ("TCS Toronto Waterfront Marathon", "MARATHON"): 2000,
    ("EDP Maratona de Lisboa", "MARATHON"): 1986,
    ("Zurich Maratón de Sevilla", "MARATHON"): 1985,
    ("Zurich Marató de Barcelona", "MARATHON"): 1978,
    ("Acea Run Rome The Marathon", "MARATHON"): 1982,
    ("Zurich Rock 'n' Roll Running Series Madrid", "MARATHON"): 2013,
    ("NN Maraton Warszawski", "MARATHON"): 1979,
    ("Irish Life Dublin Marathon", "MARATHON"): 1980,
    ("TCS Sydney Marathon presented by ASICS", "MARATHON"): 2000,
    ("Maratón de la Ciudad de México Telcel", "MARATHON"): 1983,
    ("Bangsaen42 Chonburi Marathon", "MARATHON"): 2017,
    ("Grandma's Marathon", "MARATHON"): 1977,
    ("Dick's Pittsburgh Marathon", "MARATHON"): 2009,

    # SEMI
    ("HOKA Semi de Paris", "SEMI"): 1993,
    ("Generali Berlin Half Marathon", "SEMI"): 1990,
    ("United Airlines NYC Half", "SEMI"): 2006,
    ("AJ Bell Great North Run", "SEMI"): 1981,
    ("Copenhagen Half Marathon", "SEMI"): 2015,
    ("Great Scottish Run", "SEMI"): 2006,
    ("AJ Bell Great Manchester Run", "SEMI"): 2014,
    ("AJ Bell Great Bristol Run", "SEMI"): 2001,
    ("AJ Bell Great Birmingham Run", "SEMI"): 2003,
    ("Eurospin RomaOstia Half Marathon", "SEMI"): 1975,
    ("Mitja Marato Barcelona by Brooks", "SEMI"): 2005,
    ("NN CPC Loop Den Haag - Half Marathon", "SEMI"): 1975,
    ("Generali Prague Half Marathon", "SEMI"): 2000,
    ("EDP Lisboa Meia Maratona", "SEMI"): 1991,
    ("London Landmarks Half Marathon", "SEMI"): 2018,
    ("Cardiff Half Marathon", "SEMI"): 2003,
    ("Manchester Half Marathon", "SEMI"): 2007,
    ("Royal Parks Half Marathon", "SEMI"): 2008,
    ("The Big Half", "SEMI"): 2018,
    ("Valencia Half Marathon Trinidad Alfonso Zurich", "SEMI"): 1992,
    ("Run in Lyon", "SEMI"): 2014,
    ("Medio Maraton de Sevilla", "SEMI"): 1994,
    ("TCS Toronto Waterfront Marathon", "SEMI"): 2000,
    ("Tata Mumbai Marathon", "SEMI"): 2004,
    ("Vienna City Half Marathon", "SEMI"): 1984,
    ("Warsaw Half Marathon", "SEMI"): 2010,
    ("GetPro Bath Half", "SEMI"): 1982,
    ("IU Health 500 Festival Mini-Marathon", "SEMI"): 1977,
    ("Walt Disney World Marathon Weekend", "SEMI"): 1994,
    ("Aramco Houston Half Marathon", "SEMI"): 2002,
    ("Standard Chartered Hong Kong Half Marathon", "SEMI"): 1997,
    ("Rock 'n' Roll Running Series Las Vegas", "SEMI"): 2009,
    ("Burj2Burj Half Marathon", "SEMI"): 2023,
    ("Kagawa Marugame International Half Marathon", "SEMI"): 1985,
    ("Coelmo Napoli City Half Marathon", "SEMI"): 2014,
    ("Movistar Madrid Medio Maraton", "SEMI"): 1978,
    ("Disney Princess Half Marathon", "SEMI"): 2009,
    ("St. Jude Rock 'n' Roll Running Series Washington DC", "SEMI"): 2003,
    ("NYCRUNS Brooklyn Experience Half Marathon", "SEMI"): 2017,
    ("Zurich Rock 'n' Roll Running Series Madrid", "SEMI"): 2013,
    ("21K de Montréal", "SEMI"): 2003,
    ("Göteborgsvarvet", "SEMI"): 1980,
    ("St. Jude Rock 'n' Roll Series Nashville", "SEMI"): 2000,
    ("NYRR RBC Brooklyn Half", "SEMI"): 2012,
    ("Hoka Runaway Sydney Half Marathon", "SEMI"): 2022,
    ("Rock 'n' Roll Running Series San Diego", "SEMI"): 1998,
    ("AJ Bell Great Manchester Run", "SEMI"): 2014,
    ("AJ Bell Great Bristol Run", "SEMI"): 2001,
    ("Bröllopet - The Bridge Run", "SEMI"): 2000,
    ("Media Maraton de Bogota", "SEMI"): 2000,
    ("Asics Run Melbourne", "SEMI"): 2012,
    ("Life Time Chicago Half Marathon & 5K", "SEMI"): 2000,
    ("Wizz Air Rome Half Marathon by Brooks", "SEMI"): 2014,
    ("Tokyo Legacy Half Marathon", "SEMI"): 2022,
    ("TCS Mizuno Half Marathon", "SEMI"): 2013,
    ("Nike Melbourne Marathon Festival", "SEMI"): 2009,
    ("Vedanta Delhi Half Marathon", "SEMI"): 2005,
    ("Half Marathon Munchen by Brooks", "SEMI"): 2019,
    ("Hyundai Meia Maratona", "SEMI"): 2004,
    ("Semi-Marathon de Boulogne-Billancourt", "SEMI"): 1995,
    ("Taipei Half Marathon", "SEMI"): 2017,
    ("Standard Chartered Singapore Half Marathon", "SEMI"): 2002,
    ("Bangsaen21", "SEMI"): 2017,
    ("Rock 'n' Roll Running Series Mexico City", "SEMI"): 2017,
    ("Riyadh Marathon", "SEMI"): 2022,

    # 10KM
    ("Atlanta Journal-Constitution Peachtree Road Race", "10KM"): 1970,
    ("BOLDERBoulder 10K", "10KM"): 1979,
    ("Adidas 10K Paris", "10KM"): 2007,
    ("Cooper River Bridge Run", "10KM"): 1978,
    ("Saucony London 10K", "10KM"): 2009,
    ("Cancer Research UK London Winter Run", "10KM"): 2015,
    ("10K Montmartre", "10KM"): 2024,
    ("NN CPC Loop Den Haag - 10 KM Loop", "10KM"): 1975,
    ("AJ Bell Great Manchester Run", "10KM"): 2003,
    ("AJ Bell Great Birmingham Run 10K", "10KM"): 2003,
    ("AJ Bell Great Bristol Run 10K", "10KM"): 2003,
    ("Great Scottish Run 10K", "10KM"): 1982,
    ("AJ Bell Great North 10K", "10KM"): 2013,
    ("Run in Lyon", "10KM"): 2014,
    ("Vancouver Sun Run", "10KM"): 1985,
    ("10K Valencia Ibercaja by Kiprun", "10KM"): 2019,
    ("ASICS LDNX", "10KM"): 2024,
    ("Statesman Capitol 10K", "10KM"): 1978,
    ("Ukrop's Monument Avenue 10K", "10KM"): 2000,
    ("EDP Meia Maratona de Lisboa - Vodafone 10K", "10KM"): 2010,

    # AUTRE
    ("Broad Street Run", "AUTRE"): 1980,
    ("Lilac Bloomsday Run", "AUTRE"): 1977,
    ("Bay to Breakers", "AUTRE"): 1912,
    ("Boilermaker Road Race", "AUTRE"): 1978,
    ("Manchester Road Race", "AUTRE"): 1927,
    ("Falmouth Road Race", "AUTRE"): 1973,
    ("Dam tot Damloop", "AUTRE"): 1985,
    ("City2Surf", "AUTRE"): 1971,
    ("AJ Bell Great South Run", "AUTRE"): 1990,
    ("Army Ten Miler", "AUTRE"): 1985,
    ("Gasparilla Distance Classic", "AUTRE"): 1978,
    ("Shamrock Run Portland", "AUTRE"): 1979,
    ("Credit Union Cherry Blossom", "AUTRE"): 1973,
}

wb = openpyxl.load_workbook("Suivi_Finishers_Monde_10k_-_21k_-_42k_HISTORIQUE.xlsx")
ws = wb["ALL"]
header = [c.value for c in ws[1]]
year_cols = {int(h): i + 1 for i, h in enumerate(header) if isinstance(h, (int, float))}

matched = 0
marked = 0
unmatched = []

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row[3]:
        continue
    name = str(row[3]).strip()
    dist = str(row[2]).strip()

    first_year = FIRST_EDITIONS.get((name, dist))
    if first_year is None:
        # Try with special chars
        for (fn, fd), fy in FIRST_EDITIONS.items():
            if fd == dist and (fn in name or name in fn):
                first_year = fy
                break

    if first_year is None:
        unmatched.append(f"{name} ({dist})")
        continue

    matched += 1
    for year, col_idx in year_cols.items():
        if year < first_year:
            current = ws.cell(row=row_idx, column=col_idx).value
            if current is None or str(current).strip() == "":
                ws.cell(row=row_idx, column=col_idx, value="x")
                marked += 1

wb.save("Suivi_Finishers_Monde_10k_-_21k_-_42k_HISTORIQUE.xlsx")
print(f"Matched {matched} events, marked {marked} cells with 'x'")
if unmatched:
    print(f"\nUnmatched events ({len(unmatched)}):")
    for u in unmatched:
        print(f"  - {u}")
