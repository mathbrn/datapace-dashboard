#!/usr/bin/env python3
"""
Add a new event to the Excel file in the 'ALL' sheet.

Usage:
    python add_event.py "Period" "City" "Distance" "Race Name" [YEAR COUNT ...]
    python add_event.py "Avril" "Hamburg" "MARATHON" "Haspa Marathon Hamburg" 2023 12000 2024 13500 2025 15000

Distance must be: MARATHON, SEMI, 10KM
Period must be: Janvier, Février, Mars, Avril, Mai, Juin, Juillet, Août, Septembre, Octobre, Novembre, Décembre
"""
import sys
import openpyxl
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent
EXCEL_FILE = SCRIPT_DIR / "Suivi_Finishers_Monde_10k_-_21k_-_42k_HISTORIQUE.xlsx"


def add_event(period, city, distance, race_name, year_data):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["ALL"]

    header = [cell.value for cell in ws[1]]

    # Check if race already exists
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[3]).strip() == race_name and str(row[2]).strip() == distance:
            print(f"ERREUR: '{race_name}' ({distance}) existe deja dans le fichier.")
            return False

    # Find next empty row
    next_row = ws.max_row + 1

    # Write period, city, distance, race
    ws.cell(row=next_row, column=1, value=period)
    ws.cell(row=next_row, column=2, value=city)
    ws.cell(row=next_row, column=3, value=distance)
    ws.cell(row=next_row, column=4, value=race_name)

    # Write year data
    for year, count in year_data.items():
        year_col = None
        for i, h in enumerate(header):
            if isinstance(h, (int, float)) and int(h) == int(year):
                year_col = i + 1
                break
        if year_col:
            ws.cell(row=next_row, column=year_col, value=int(count))

    wb.save(EXCEL_FILE)
    years_str = ", ".join(f"{y}:{c}" for y, c in year_data.items())
    print(f"[AJOUTE] {race_name} ({distance}) - {city} - {period} | {years_str}")
    return True


if __name__ == "__main__":
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)

    period = sys.argv[1]
    city = sys.argv[2]
    distance = sys.argv[3]
    race_name = sys.argv[4]

    if distance not in ("MARATHON", "SEMI", "10KM"):
        print(f"ERREUR: Distance '{distance}' invalide.")
        sys.exit(1)

    year_data = {}
    args = sys.argv[5:]
    for i in range(0, len(args), 2):
        if i + 1 < len(args):
            year_data[int(args[i])] = int(args[i + 1])

    add_event(period, city, distance, race_name, year_data)
