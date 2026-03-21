import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# All collected data: (Course, Distance, Année, Temps Homme, Temps Femme)
data = []

# ============================================================
# MARATHONS (42K)
# ============================================================

marathon_data = [
    # ============================================================
    # HISTORICAL DATA (2015-2022)
    # ============================================================
    # BMW Berlin Marathon
    ("BMW Berlin Marathon", "42K", 2015, "2:04:00", "2:19:25"),
    ("BMW Berlin Marathon", "42K", 2016, "2:03:03", "N/A"),
    ("BMW Berlin Marathon", "42K", 2017, "2:03:32", "2:20:53"),
    ("BMW Berlin Marathon", "42K", 2018, "2:01:39", "2:18:11"),
    ("BMW Berlin Marathon", "42K", 2019, "2:01:41", "N/A"),
    ("BMW Berlin Marathon", "42K", 2021, "2:05:45", "N/A"),
    ("BMW Berlin Marathon", "42K", 2022, "2:01:09", "N/A"),
    # Bank of America Chicago Marathon
    ("Bank of America Chicago Marathon", "42K", 2015, "2:09:25", "2:23:33"),
    ("Bank of America Chicago Marathon", "42K", 2016, "2:11:23", "2:21:32"),
    ("Bank of America Chicago Marathon", "42K", 2017, "2:09:20", "2:18:31"),
    ("Bank of America Chicago Marathon", "42K", 2018, "2:05:11", "2:18:35"),
    ("Bank of America Chicago Marathon", "42K", 2019, "2:05:45", "2:14:04"),
    ("Bank of America Chicago Marathon", "42K", 2021, "2:06:12", "2:22:31"),
    ("Bank of America Chicago Marathon", "42K", 2022, "2:04:24", "2:14:18"),
    # TCS London Marathon
    ("TCS London Marathon", "42K", 2015, "2:04:42", "N/A"),
    ("TCS London Marathon", "42K", 2016, "2:03:05", "N/A"),
    ("TCS London Marathon", "42K", 2017, "2:05:48", "2:17:01"),
    ("TCS London Marathon", "42K", 2018, "2:04:17", "2:18:31"),
    ("TCS London Marathon", "42K", 2019, "2:02:37", "N/A"),
    ("TCS London Marathon", "42K", 2020, "2:05:41", "2:18:58"),
    ("TCS London Marathon", "42K", 2021, "2:04:01", "N/A"),
    ("TCS London Marathon", "42K", 2022, "2:04:39", "2:17:26"),
    # TCS New York City Marathon
    ("TCS New York City Marathon", "42K", 2015, "2:10:34", "2:24:25"),
    ("TCS New York City Marathon", "42K", 2016, "2:07:51", "2:24:26"),
    ("TCS New York City Marathon", "42K", 2017, "2:10:53", "N/A"),
    ("TCS New York City Marathon", "42K", 2018, "2:05:59", "2:22:48"),
    ("TCS New York City Marathon", "42K", 2019, "2:08:13", "2:22:38"),
    ("TCS New York City Marathon", "42K", 2021, "2:08:22", "2:22:39"),
    ("TCS New York City Marathon", "42K", 2022, "2:08:41", "2:23:23"),
    # Boston Marathon
    ("Boston Marathon", "42K", 2015, "2:09:17", "2:24:55"),
    ("Boston Marathon", "42K", 2016, "2:12:45", "2:29:19"),
    ("Boston Marathon", "42K", 2017, "2:09:37", "2:21:52"),
    ("Boston Marathon", "42K", 2018, "2:15:58", "2:39:54"),
    ("Boston Marathon", "42K", 2019, "2:07:57", "2:23:31"),
    ("Boston Marathon", "42K", 2021, "2:09:51", "N/A"),
    ("Boston Marathon", "42K", 2022, "2:06:51", "2:21:01"),
    # Tokyo Marathon
    ("Tokyo Marathon", "42K", 2015, "2:06:00", "2:23:15"),
    ("Tokyo Marathon", "42K", 2016, "2:06:56", "2:21:27"),
    ("Tokyo Marathon", "42K", 2017, "2:03:58", "N/A"),
    ("Tokyo Marathon", "42K", 2018, "2:05:30", "2:19:51"),
    ("Tokyo Marathon", "42K", 2019, "2:04:15", "N/A"),
    ("Tokyo Marathon", "42K", 2020, "2:04:15", "2:17:45"),
    ("Tokyo Marathon", "42K", 2021, "2:02:40", "2:16:02"),
    # Schneider Electric Marathon de Paris
    ("Schneider Electric Marathon de Paris", "42K", 2017, "2:06:10", "N/A"),
    ("Schneider Electric Marathon de Paris", "42K", 2018, "2:06:25", "2:22:56"),
    ("Schneider Electric Marathon de Paris", "42K", 2019, "2:07:05", "2:22:47"),
    ("Schneider Electric Marathon de Paris", "42K", 2022, "2:05:00", "2:19:48"),
    # NN Marathon Rotterdam
    ("NN Marathon Rotterdam", "42K", 2015, "2:06:47", "2:26:30"),
    ("NN Marathon Rotterdam", "42K", 2016, "2:06:11", "2:26:15"),
    ("NN Marathon Rotterdam", "42K", 2017, "2:06:04", "2:24:18"),
    ("NN Marathon Rotterdam", "42K", 2018, "2:04:11", "2:22:55"),
    ("NN Marathon Rotterdam", "42K", 2021, "2:03:36", "N/A"),
    ("NN Marathon Rotterdam", "42K", 2022, "2:04:45", "N/A"),
    # TCS Amsterdam Marathon
    ("TCS Amsterdam Marathon", "42K", 2015, "2:06:19", "2:24:11"),
    ("TCS Amsterdam Marathon", "42K", 2016, "2:05:20", "2:23:20"),
    ("TCS Amsterdam Marathon", "42K", 2017, "2:05:09", "2:21:53"),
    ("TCS Amsterdam Marathon", "42K", 2018, "2:04:06", "N/A"),
    ("TCS Amsterdam Marathon", "42K", 2021, "2:03:38", "2:17:57"),
    ("TCS Amsterdam Marathon", "42K", 2022, "N/A", "2:17:20"),
    # Valencia Marathon Trinidad Alfonso Zurich
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2015, "2:06:13", "N/A"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2020, "2:03:00", "N/A"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2022, "2:01:53", "2:14:57"),
    # Irish Life Dublin Marathon
    ("Irish Life Dublin Marathon", "42K", 2015, "2:14:01", "2:30:00"),
    ("Irish Life Dublin Marathon", "42K", 2016, "2:12:18", "2:32:32"),
    ("Irish Life Dublin Marathon", "42K", 2017, "2:15:52", "2:28:57"),
    ("Irish Life Dublin Marathon", "42K", 2018, "2:13:23", "2:33:48"),
    ("Irish Life Dublin Marathon", "42K", 2019, "2:08:06", "2:27:48"),
    ("Irish Life Dublin Marathon", "42K", 2022, "2:11:30", "2:28:32"),
    # Vienna City Marathon
    ("Vienna City Marathon", "42K", 2015, "2:07:31", "2:30:09"),
    ("Vienna City Marathon", "42K", 2017, "2:08:40", "2:24:20"),
    ("Vienna City Marathon", "42K", 2019, "2:06:56", "2:22:12"),
    ("Vienna City Marathon", "42K", 2021, "2:09:25", "2:24:29"),
    ("Vienna City Marathon", "42K", 2022, "2:06:53", "2:20:59"),
    # Honolulu Marathon
    ("Honolulu Marathon", "42K", 2015, "2:11:43", "2:28:34"),
    ("Honolulu Marathon", "42K", 2016, "2:09:37", "2:31:10"),
    ("Honolulu Marathon", "42K", 2017, "2:08:27", "2:22:15"),
    ("Honolulu Marathon", "42K", 2018, "2:09:01", "2:36:22"),
    ("Honolulu Marathon", "42K", 2019, "2:07:59", "2:31:09"),
    ("Honolulu Marathon", "42K", 2021, "2:14:30", "2:41:24"),
    ("Honolulu Marathon", "42K", 2022, "2:14:40", "2:30:58"),
    # ============================================================
    # RECENT DATA (2023-2026)
    # ============================================================
    # Chevron Houston Marathon
    ("Chevron Houston Marathon", "42K", 2023, "2:10:26", "2:19:24"),
    ("Chevron Houston Marathon", "42K", 2024, "2:06:39", "2:19:33"),
    ("Chevron Houston Marathon", "42K", 2025, "2:08:17", "2:20:54"),
    ("Chevron Houston Marathon", "42K", 2026, "2:05:45", "2:24:17"),
    # Tata Mumbai Marathon
    ("Tata Mumbai Marathon", "42K", 2023, "2:07:32", "2:24:15"),
    ("Tata Mumbai Marathon", "42K", 2024, "2:07:50", "2:26:06"),
    ("Tata Mumbai Marathon", "42K", 2025, "2:11:44", "2:24:56"),
    ("Tata Mumbai Marathon", "42K", 2026, "2:09:55", "2:25:13"),
    # Zurich Maratón de Sevilla
    ("Zurich Maratón de Sevilla", "42K", 2023, "2:04:59", "2:20:29"),
    ("Zurich Maratón de Sevilla", "42K", 2024, "2:03:27", "2:22:13"),
    ("Zurich Maratón de Sevilla", "42K", 2025, "2:05:15", "2:22:17"),
    ("Zurich Maratón de Sevilla", "42K", 2026, "2:03:59", "2:20:39"),
    # Seoul Marathon
    ("Seoul Marathon", "42K", 2023, "2:05:27", "2:28:32"),
    ("Seoul Marathon", "42K", 2024, "2:06:08", "2:21:32"),
    ("Seoul Marathon", "42K", 2025, "2:05:42", "2:21:36"),
    # Tokyo Marathon
    ("Tokyo Marathon", "42K", 2023, "2:05:22", "2:16:28"),
    ("Tokyo Marathon", "42K", 2024, "2:02:16", "2:15:55"),
    ("Tokyo Marathon", "42K", 2025, "2:03:23", "2:16:31"),
    ("Tokyo Marathon", "42K", 2026, "2:03:37", "2:14:29"),
    # Zurich Marató de Barcelona
    ("Zurich Marató de Barcelona", "42K", 2023, "2:05:06", "2:19:44"),
    ("Zurich Marató de Barcelona", "42K", 2024, "2:05:01", "2:19:52"),
    ("Zurich Marató de Barcelona", "42K", 2025, "2:04:13", "2:19:33"),
    ("Zurich Marató de Barcelona", "42K", 2026, "2:04:57", "2:10:53"),
    # Acea Run Rome The Marathon
    ("Acea Run Rome The Marathon", "42K", 2023, "2:07:43", "2:23:01"),
    ("Acea Run Rome The Marathon", "42K", 2024, "2:06:23", "2:24:35"),
    ("Acea Run Rome The Marathon", "42K", 2025, "2:07:35", "2:26:16"),
    # ASICS Los Angeles Marathon
    ("ASICS Los Angeles Marathon", "42K", 2023, "2:13:13", "2:31:00"),
    ("ASICS Los Angeles Marathon", "42K", 2024, "2:11:00", "2:25:29"),
    ("ASICS Los Angeles Marathon", "42K", 2025, "2:07:56", "2:30:16"),
    ("ASICS Los Angeles Marathon", "42K", 2026, "2:11:18", "2:25:20"),
    # Boston Marathon
    ("Boston Marathon", "42K", 2023, "2:05:54", "2:21:38"),
    ("Boston Marathon", "42K", 2024, "2:06:17", "2:22:37"),
    ("Boston Marathon", "42K", 2025, "2:04:45", "2:17:22"),
    # TCS London Marathon
    ("TCS London Marathon", "42K", 2023, "2:01:25", "2:18:33"),
    ("TCS London Marathon", "42K", 2024, "2:04:01", "2:16:16"),
    ("TCS London Marathon", "42K", 2025, "2:02:27", "2:15:50"),
    # Zurich Rock 'n' Roll Running Series Madrid
    ("Zurich Rock 'n' Roll Running Series Madrid", "42K", 2023, "2:10:29", "2:26:31"),
    ("Zurich Rock 'n' Roll Running Series Madrid", "42K", 2024, "2:08:05", "2:26:19"),
    ("Zurich Rock 'n' Roll Running Series Madrid", "42K", 2025, "2:09:11", "2:25:55"),
    # Adidas Manchester Marathon
    ("Adidas Manchester Marathon", "42K", 2023, "2:16:27", "2:31:26"),
    ("Adidas Manchester Marathon", "42K", 2024, "2:16:29", "2:37:14"),
    ("Adidas Manchester Marathon", "42K", 2025, "2:16:56", "2:34:53"),
    # Schneider Electric Marathon de Paris
    ("Schneider Electric Marathon de Paris", "42K", 2023, "2:07:15", "2:23:19"),
    ("Schneider Electric Marathon de Paris", "42K", 2024, "2:05:33", "2:20:45"),
    ("Schneider Electric Marathon de Paris", "42K", 2025, "2:05:25", "2:20:45"),
    # NN Marathon Rotterdam
    ("NN Marathon Rotterdam", "42K", 2023, "2:03:47", "2:20:31"),
    ("NN Marathon Rotterdam", "42K", 2024, "2:04:45", "2:19:30"),
    ("NN Marathon Rotterdam", "42K", 2025, "2:04:33", "2:21:15"),
    # Vienna City Marathon
    ("Vienna City Marathon", "42K", 2023, "2:05:08", "2:24:12"),
    ("Vienna City Marathon", "42K", 2024, "2:06:35", "2:24:08"),
    ("Vienna City Marathon", "42K", 2025, "2:08:28", "2:24:14"),
    # Copenhagen Marathon
    ("Copenhagen Marathon", "42K", 2023, "2:09:12", "2:23:14"),
    ("Copenhagen Marathon", "42K", 2024, "2:09:11", "2:23:19"),
    ("Copenhagen Marathon", "42K", 2025, "2:09:09", "2:23:19"),
    # Prague International Marathon
    ("Prague International Marathon", "42K", 2023, "2:05:09", "2:20:42"),
    ("Prague International Marathon", "42K", 2024, "2:08:43", "2:23:41"),
    ("Prague International Marathon", "42K", 2025, "2:05:14", "2:20:55"),
    # Sanlam Cape Town Marathon
    ("Sanlam Cape Town Marathon", "42K", 2023, "2:11:28", "2:24:17"),
    ("Sanlam Cape Town Marathon", "42K", 2024, "2:08:16", "2:22:22"),
    ("Sanlam Cape Town Marathon", "42K", 2025, "Annulé", "Annulé"),
    # Maratón de la Ciudad de México Telcel
    ("Maratón de la Ciudad de México Telcel", "42K", 2023, "2:08:23", "2:27:17"),
    ("Maratón de la Ciudad de México Telcel", "42K", 2024, "2:10:36", "2:29:19"),
    ("Maratón de la Ciudad de México Telcel", "42K", 2025, "2:11:14", "2:23:22"),
    # TCS Sydney Marathon
    ("TCS Sydney Marathon presented by ASICS", "42K", 2023, "2:08:20", "2:26:47"),
    ("TCS Sydney Marathon presented by ASICS", "42K", 2024, "2:06:17", "2:21:40"),
    ("TCS Sydney Marathon presented by ASICS", "42K", 2025, "2:06:06", "2:18:22"),
    # BMW Berlin Marathon
    ("BMW Berlin Marathon", "42K", 2023, "2:02:42", "2:11:53"),
    ("BMW Berlin Marathon", "42K", 2024, "2:03:17", "2:16:42"),
    ("BMW Berlin Marathon", "42K", 2025, "2:02:16", "2:21:05"),
    # NN Maraton Warszawski
    ("NN Maraton Warszawski", "42K", 2023, "2:15:29", "2:34:41"),
    ("NN Maraton Warszawski", "42K", 2024, "2:10:43", "2:31:24"),
    ("NN Maraton Warszawski", "42K", 2025, "2:11:21", "2:33:41"),
    # TCS Amsterdam Marathon
    ("TCS Amsterdam Marathon", "42K", 2023, "2:04:18", "2:18:21"),
    ("TCS Amsterdam Marathon", "42K", 2024, "2:05:38", "2:16:52"),
    ("TCS Amsterdam Marathon", "42K", 2025, "2:03:31", "2:17:37"),
    # Bank of America Chicago Marathon
    ("Bank of America Chicago Marathon", "42K", 2023, "2:00:35", "2:13:44"),
    ("Bank of America Chicago Marathon", "42K", 2024, "2:02:43", "2:09:56"),
    ("Bank of America Chicago Marathon", "42K", 2025, "2:02:21", "2:14:56"),
    # EDP Maratona de Lisboa
    ("EDP Maratona de Lisboa", "42K", 2023, "N/A", "N/A"),
    ("EDP Maratona de Lisboa", "42K", 2024, "N/A", "N/A"),
    ("EDP Maratona de Lisboa", "42K", 2025, "2:05:43", "2:24:17"),
    # TCS Toronto Waterfront Marathon
    ("TCS Toronto Waterfront Marathon", "42K", 2023, "2:09:20", "2:23:11"),
    ("TCS Toronto Waterfront Marathon", "42K", 2024, "2:07:16", "2:20:44"),
    ("TCS Toronto Waterfront Marathon", "42K", 2025, "2:08:05", "2:21:04"),
    # Irish Life Dublin Marathon
    ("Irish Life Dublin Marathon", "42K", 2023, "2:06:52", "2:26:22"),
    ("Irish Life Dublin Marathon", "42K", 2024, "2:08:47", "2:24:13"),
    ("Irish Life Dublin Marathon", "42K", 2025, "2:08:51", "2:26:28"),
    # Marine Corps Marathon
    ("Marine Corps Marathon", "42K", 2023, "2:25:56", "2:50:49"),
    ("Marine Corps Marathon", "42K", 2024, "2:25:06", "2:39:36"),
    ("Marine Corps Marathon", "42K", 2025, "2:18:51", "2:34:08"),
    # Bangsaen42 Chonburi Marathon
    ("Bangsaen42 Chonburi Marathon", "42K", 2023, "N/A", "N/A"),
    ("Bangsaen42 Chonburi Marathon", "42K", 2024, "N/A", "N/A"),
    ("Bangsaen42 Chonburi Marathon", "42K", 2025, "2:17:53", "N/A"),
    # TCS New York City Marathon
    ("TCS New York City Marathon", "42K", 2023, "2:04:58", "2:27:23"),
    ("TCS New York City Marathon", "42K", 2024, "2:07:39", "2:24:35"),
    ("TCS New York City Marathon", "42K", 2025, "2:08:09", "2:19:51"),
    # Standard Chartered Singapore Marathon
    ("Standard Chartered Singapore Marathon", "42K", 2024, "2:16:06", "2:39:04"),
    ("Standard Chartered Singapore Marathon", "42K", 2025, "2:15:40", "2:41:24"),
    # Taipei Marathon
    ("Taipei Marathon", "42K", 2023, "2:11:05", "2:27:14"),
    ("Taipei Marathon", "42K", 2024, "2:11:41", "2:32:47"),
    ("Taipei Marathon", "42K", 2025, "N/A", "N/A"),  # Course en décembre 2025, pas encore de données fiables
    # Valencia Marathon Trinidad Alfonso Zurich
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2023, "2:01:48", "2:15:51"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2024, "2:02:05", "2:16:49"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2025, "2:02:24", "2:14:00"),
    # Honolulu Marathon
    ("Honolulu Marathon", "42K", 2023, "2:15:42", "2:33:01"),
    ("Honolulu Marathon", "42K", 2024, "2:11:59", "2:31:14"),
    ("Honolulu Marathon", "42K", 2025, "2:13:38", "2:30:43"),
    # Standard Chartered KL Marathon
    ("Standard Chartered KL Marathon", "42K", 2025, "2:17:28", "2:41:36"),
    # Standard Chartered Hong Kong Marathon
    ("Standard Chartered Hong Kong Marathon", "42K", 2026, "N/A", "N/A"),
    ("Standard Chartered Hong Kong Marathon", "42K", 2019, "2:09:20", "2:26:13"),
    # ── AGENT 1: WMM 2000-2014 ──
    # BMW Berlin Marathon 2000-2014
    ("BMW Berlin Marathon", "42K", 2000, "2:07:42", "2:26:20"),
    ("BMW Berlin Marathon", "42K", 2001, "2:08:47", "2:26:05"),
    ("BMW Berlin Marathon", "42K", 2002, "2:06:47", "2:21:49"),
    ("BMW Berlin Marathon", "42K", 2003, "2:04:55", "2:19:46"),
    ("BMW Berlin Marathon", "42K", 2004, "2:06:44", "2:19:41"),
    ("BMW Berlin Marathon", "42K", 2005, "2:07:41", "2:19:59"),
    ("BMW Berlin Marathon", "42K", 2006, "2:05:56", "2:21:34"),
    ("BMW Berlin Marathon", "42K", 2007, "2:04:26", "2:23:17"),
    ("BMW Berlin Marathon", "42K", 2008, "2:03:59", "2:19:19"),
    ("BMW Berlin Marathon", "42K", 2009, "2:06:08", "2:25:10"),
    ("BMW Berlin Marathon", "42K", 2010, "2:05:08", "2:23:08"),
    ("BMW Berlin Marathon", "42K", 2011, "2:03:38", "2:19:12"),
    ("BMW Berlin Marathon", "42K", 2012, "2:04:15", "2:19:34"),
    ("BMW Berlin Marathon", "42K", 2013, "2:03:23", "2:21:13"),
    ("BMW Berlin Marathon", "42K", 2014, "2:02:57", "2:20:18"),
    # Bank of America Chicago Marathon 2000-2014
    ("Bank of America Chicago Marathon", "42K", 2000, "2:07:01", "2:21:33"),
    ("Bank of America Chicago Marathon", "42K", 2001, "2:08:52", "2:18:47"),
    ("Bank of America Chicago Marathon", "42K", 2002, "2:05:56", "2:18:18"),
    ("Bank of America Chicago Marathon", "42K", 2003, "2:05:50", "2:21:18"),
    ("Bank of America Chicago Marathon", "42K", 2004, "2:06:16", "2:23:45"),
    ("Bank of America Chicago Marathon", "42K", 2005, "2:07:02", "2:21:25"),
    ("Bank of America Chicago Marathon", "42K", 2006, "2:07:35", "2:20:42"),
    ("Bank of America Chicago Marathon", "42K", 2007, "2:11:11", "2:33:49"),
    ("Bank of America Chicago Marathon", "42K", 2008, "2:06:25", "2:25:28"),
    ("Bank of America Chicago Marathon", "42K", 2009, "2:05:41", "2:25:56"),
    ("Bank of America Chicago Marathon", "42K", 2010, "2:06:24", "2:20:25"),
    ("Bank of America Chicago Marathon", "42K", 2011, "2:05:37", "2:25:05"),
    ("Bank of America Chicago Marathon", "42K", 2012, "2:04:38", "2:19:09"),
    ("Bank of America Chicago Marathon", "42K", 2013, "2:03:45", "2:22:44"),
    ("Bank of America Chicago Marathon", "42K", 2014, "2:04:11", "2:24:31"),
    # TCS New York City Marathon 2000-2014
    ("TCS New York City Marathon", "42K", 2000, "2:10:09", "2:25:45"),
    ("TCS New York City Marathon", "42K", 2001, "2:07:43", "2:24:21"),
    ("TCS New York City Marathon", "42K", 2002, "2:08:07", "2:22:31"),
    ("TCS New York City Marathon", "42K", 2003, "2:10:30", "2:22:31"),
    ("TCS New York City Marathon", "42K", 2004, "2:09:28", "2:23:10"),
    ("TCS New York City Marathon", "42K", 2005, "2:09:30", "2:24:41"),
    ("TCS New York City Marathon", "42K", 2006, "2:09:58", "2:25:05"),
    ("TCS New York City Marathon", "42K", 2007, "2:09:04", "2:23:09"),
    ("TCS New York City Marathon", "42K", 2008, "2:08:43", "2:23:56"),
    ("TCS New York City Marathon", "42K", 2009, "2:09:15", "2:25:53"),
    ("TCS New York City Marathon", "42K", 2010, "2:08:14", "2:28:20"),
    ("TCS New York City Marathon", "42K", 2011, "2:05:06", "2:23:15"),
    ("TCS New York City Marathon", "42K", 2013, "2:08:24", "2:25:07"),
    ("TCS New York City Marathon", "42K", 2014, "2:10:59", "2:25:07"),
    # Boston Marathon 2000-2014
    ("Boston Marathon", "42K", 2000, "2:09:47", "2:26:11"),
    ("Boston Marathon", "42K", 2001, "2:09:43", "2:23:53"),
    ("Boston Marathon", "42K", 2002, "2:09:02", "2:20:43"),
    ("Boston Marathon", "42K", 2003, "2:10:11", "2:25:20"),
    ("Boston Marathon", "42K", 2004, "2:10:37", "2:24:27"),
    ("Boston Marathon", "42K", 2005, "2:11:45", "2:25:13"),
    ("Boston Marathon", "42K", 2006, "2:07:14", "2:23:38"),
    ("Boston Marathon", "42K", 2007, "2:14:13", "2:29:18"),
    ("Boston Marathon", "42K", 2008, "2:07:46", "2:25:25"),
    ("Boston Marathon", "42K", 2009, "2:08:42", "2:32:16"),
    ("Boston Marathon", "42K", 2010, "2:05:52", "2:26:11"),
    ("Boston Marathon", "42K", 2011, "2:03:02", "2:22:36"),
    ("Boston Marathon", "42K", 2012, "2:12:40", "2:31:50"),
    ("Boston Marathon", "42K", 2013, "2:10:22", "2:26:25"),
    ("Boston Marathon", "42K", 2014, "2:08:37", "2:19:59"),
    # TCS London Marathon 2000-2014
    ("TCS London Marathon", "42K", 2000, "2:06:36", "2:24:33"),
    ("TCS London Marathon", "42K", 2001, "2:07:11", "2:23:56"),
    ("TCS London Marathon", "42K", 2002, "2:05:38", "2:18:56"),
    ("TCS London Marathon", "42K", 2003, "2:07:56", "2:15:25"),
    ("TCS London Marathon", "42K", 2004, "2:06:18", "2:22:35"),
    ("TCS London Marathon", "42K", 2005, "2:07:44", "2:17:42"),
    ("TCS London Marathon", "42K", 2006, "2:06:39", "2:20:04"),
    ("TCS London Marathon", "42K", 2007, "2:07:41", "2:20:13"),
    ("TCS London Marathon", "42K", 2008, "2:05:15", "2:24:14"),
    ("TCS London Marathon", "42K", 2009, "2:05:10", "2:23:12"),
    ("TCS London Marathon", "42K", 2010, "2:05:19", "2:22:00"),
    ("TCS London Marathon", "42K", 2011, "2:04:40", "2:19:36"),
    ("TCS London Marathon", "42K", 2012, "2:04:44", "2:18:37"),
    ("TCS London Marathon", "42K", 2013, "2:04:29", "2:20:15"),
    ("TCS London Marathon", "42K", 2014, "2:04:29", "2:20:21"),
    # Tokyo Marathon 2007-2014
    ("Tokyo Marathon", "42K", 2007, "2:09:45", "2:35:28"),
    ("Tokyo Marathon", "42K", 2008, "2:10:57", "2:25:51"),
    ("Tokyo Marathon", "42K", 2009, "2:10:25", "2:25:38"),
    ("Tokyo Marathon", "42K", 2010, "2:08:12", "2:25:27"),
    ("Tokyo Marathon", "42K", 2011, "2:07:32", "2:26:49"),
    ("Tokyo Marathon", "42K", 2012, "2:07:37", "2:25:28"),
    ("Tokyo Marathon", "42K", 2013, "2:06:50", "2:25:34"),
    ("Tokyo Marathon", "42K", 2014, "2:05:42", "2:22:23"),
    # Schneider Electric Marathon de Paris 2000-2016
    ("Schneider Electric Marathon de Paris", "42K", 2000, "2:08:49", "2:28:34"),
    ("Schneider Electric Marathon de Paris", "42K", 2001, "2:09:40", "2:24:33"),
    ("Schneider Electric Marathon de Paris", "42K", 2002, "2:08:18", "2:24:34"),
    ("Schneider Electric Marathon de Paris", "42K", 2003, "2:06:33", "2:24:03"),
    ("Schneider Electric Marathon de Paris", "42K", 2004, "2:08:56", "2:27:01"),
    ("Schneider Electric Marathon de Paris", "42K", 2005, "2:08:02", "2:24:55"),
    ("Schneider Electric Marathon de Paris", "42K", 2006, "2:08:03", "2:29:22"),
    ("Schneider Electric Marathon de Paris", "42K", 2007, "2:07:17", "2:27:03"),
    ("Schneider Electric Marathon de Paris", "42K", 2008, "2:06:40", "2:25:15"),
    ("Schneider Electric Marathon de Paris", "42K", 2009, "2:05:47", "2:25:08"),
    ("Schneider Electric Marathon de Paris", "42K", 2010, "2:06:33", "2:22:34"),
    ("Schneider Electric Marathon de Paris", "42K", 2011, "2:06:29", "2:22:19"),
    ("Schneider Electric Marathon de Paris", "42K", 2012, "2:06:03", "2:22:28"),
    ("Schneider Electric Marathon de Paris", "42K", 2013, "2:06:37", "2:25:55"),
    ("Schneider Electric Marathon de Paris", "42K", 2014, "2:05:11", "2:22:48"),
    ("Schneider Electric Marathon de Paris", "42K", 2015, "2:05:49", "2:25:29"),
    ("Schneider Electric Marathon de Paris", "42K", 2016, "2:07:32", "2:25:26"),
    # NN Marathon Rotterdam 2000-2014
    ("NN Marathon Rotterdam", "42K", 2000, "2:07:26", "2:28:28"),
    ("NN Marathon Rotterdam", "42K", 2001, "2:06:50", "2:25:29"),
    ("NN Marathon Rotterdam", "42K", 2002, "2:06:14", "2:23:43"),
    ("NN Marathon Rotterdam", "42K", 2003, "2:08:23", "2:23:38"),
    ("NN Marathon Rotterdam", "42K", 2004, "2:06:14", "2:27:35"),
    ("NN Marathon Rotterdam", "42K", 2005, "2:07:26", "2:23:46"),
    ("NN Marathon Rotterdam", "42K", 2006, "2:06:58", "2:27:28"),
    ("NN Marathon Rotterdam", "42K", 2007, "2:07:33", "2:29:49"),
    ("NN Marathon Rotterdam", "42K", 2008, "2:04:53", "2:26:01"),
    ("NN Marathon Rotterdam", "42K", 2009, "2:05:36", "2:25:33"),
    ("NN Marathon Rotterdam", "42K", 2010, "2:04:48", "2:28:20"),
    ("NN Marathon Rotterdam", "42K", 2011, "2:04:27", "2:23:07"),
    ("NN Marathon Rotterdam", "42K", 2012, "2:04:45", "2:22:09"),
    ("NN Marathon Rotterdam", "42K", 2013, "2:06:22", "2:24:31"),
    ("NN Marathon Rotterdam", "42K", 2014, "2:05:00", "2:23:43"),
    # TCS Amsterdam Marathon 2000-2014
    ("TCS Amsterdam Marathon", "42K", 2000, "2:08:47", "2:28:42"),
    ("TCS Amsterdam Marathon", "42K", 2001, "2:08:22", "2:26:04"),
    ("TCS Amsterdam Marathon", "42K", 2002, "2:07:45", "2:23:54"),
    ("TCS Amsterdam Marathon", "42K", 2003, "2:06:38", "2:27:08"),
    ("TCS Amsterdam Marathon", "42K", 2004, "2:06:20", "2:28:39"),
    ("TCS Amsterdam Marathon", "42K", 2005, "2:09:27", "2:28:18"),
    ("TCS Amsterdam Marathon", "42K", 2006, "2:08:41", "2:27:37"),
    ("TCS Amsterdam Marathon", "42K", 2007, "2:07:57", "2:28:41"),
    ("TCS Amsterdam Marathon", "42K", 2008, "2:06:13", "2:29:02"),
    ("TCS Amsterdam Marathon", "42K", 2009, "2:06:00", "2:25:46"),
    ("TCS Amsterdam Marathon", "42K", 2010, "2:05:44", "2:28:28"),
    ("TCS Amsterdam Marathon", "42K", 2011, "2:05:16", "2:24:34"),
    ("TCS Amsterdam Marathon", "42K", 2012, "2:05:36", "2:21:09"),
    ("TCS Amsterdam Marathon", "42K", 2013, "2:05:36", "2:24:27"),
    ("TCS Amsterdam Marathon", "42K", 2014, "2:05:36", "2:23:27"),
    # Honolulu Marathon 2000-2014
    ("Honolulu Marathon", "42K", 2000, "2:15:13", "2:30:01"),
    ("Honolulu Marathon", "42K", 2001, "2:12:29", "2:32:57"),
    ("Honolulu Marathon", "42K", 2002, "2:14:28", "2:32:22"),
    ("Honolulu Marathon", "42K", 2003, "2:12:29", "2:30:26"),
    ("Honolulu Marathon", "42K", 2004, "2:13:35", "2:30:22"),
    ("Honolulu Marathon", "42K", 2005, "2:12:00", "2:33:22"),
    ("Honolulu Marathon", "42K", 2006, "2:13:13", "2:31:40"),
    ("Honolulu Marathon", "42K", 2007, "2:17:24", "2:31:01"),
    ("Honolulu Marathon", "42K", 2008, "2:12:31", "2:30:33"),
    ("Honolulu Marathon", "42K", 2009, "2:14:53", "2:29:23"),
    ("Honolulu Marathon", "42K", 2010, "2:12:05", "2:29:32"),
    ("Honolulu Marathon", "42K", 2011, "2:13:35", "2:32:13"),
    ("Honolulu Marathon", "42K", 2012, "2:12:34", "2:31:41"),
    ("Honolulu Marathon", "42K", 2013, "2:11:42", "2:30:38"),
    ("Honolulu Marathon", "42K", 2014, "2:13:07", "2:31:17"),
    # ── AGENT 4: Other marathons 2015-2025 ──
    # Athens Marathon
    ("Athens Marathon", "42K", 2015, "2:21:22", "2:52:06"),
    ("Athens Marathon", "42K", 2016, "2:12:49", "2:38:13"),
    ("Athens Marathon", "42K", 2017, "2:12:17", "2:34:18"),
    ("Athens Marathon", "42K", 2018, "2:10:56", "2:36:46"),
    ("Athens Marathon", "42K", 2019, "2:16:34", "2:45:50"),
    ("Athens Marathon", "42K", 2021, "2:16:49", "2:41:30"),
    ("Athens Marathon", "42K", 2022, "2:23:44", "2:46:01"),
    ("Athens Marathon", "42K", 2023, "2:10:34", "2:31:52"),
    ("Athens Marathon", "42K", 2024, "2:18:56", "2:40:19"),
    ("Athens Marathon", "42K", 2025, "2:20:10", "2:39:27"),
    # Istanbul Marathon
    ("Istanbul Marathon", "42K", 2015, "2:11:17", "2:31:58"),
    ("Istanbul Marathon", "42K", 2016, "2:13:36", "2:28:23"),
    ("Istanbul Marathon", "42K", 2017, "2:11:22", "2:22:36"),
    ("Istanbul Marathon", "42K", 2018, "2:09:57", "2:18:35"),
    ("Istanbul Marathon", "42K", 2019, "2:09:44", "2:23:38"),
    ("Istanbul Marathon", "42K", 2021, "2:10:57", "2:24:15"),
    ("Istanbul Marathon", "42K", 2022, "2:09:35", "2:20:43"),
    ("Istanbul Marathon", "42K", 2023, "2:10:35", "2:27:09"),
    ("Istanbul Marathon", "42K", 2024, "2:11:40", "2:24:45"),
    ("Istanbul Marathon", "42K", 2025, "2:10:12", "2:26:19"),
    # Stockholm Marathon
    ("Stockholm Marathon", "42K", 2015, "2:18:22", "2:34:14"),
    ("Stockholm Marathon", "42K", 2016, "2:10:58", "2:31:46"),
    ("Stockholm Marathon", "42K", 2017, "2:11:36", "2:35:45"),
    ("Stockholm Marathon", "42K", 2018, "2:13:30", "2:40:28"),
    ("Stockholm Marathon", "42K", 2019, "2:10:10", "2:33:38"),
    ("Stockholm Marathon", "42K", 2021, "2:12:24", "2:29:03"),
    ("Stockholm Marathon", "42K", 2022, "2:11:08", "2:31:48"),
    ("Stockholm Marathon", "42K", 2023, "2:10:32", "2:30:44"),
    ("Stockholm Marathon", "42K", 2024, "2:14:17", "2:31:46"),
    ("Stockholm Marathon", "42K", 2025, "2:11:34", "2:30:38"),
    # Mainova Frankfurt Marathon
    ("Mainova Frankfurt Marathon", "42K", 2015, "2:06:26", "2:23:12"),
    ("Mainova Frankfurt Marathon", "42K", 2016, "2:06:48", "2:25:27"),
    ("Mainova Frankfurt Marathon", "42K", 2017, "2:05:50", "2:23:35"),
    ("Mainova Frankfurt Marathon", "42K", 2018, "2:06:37", "2:20:36"),
    ("Mainova Frankfurt Marathon", "42K", 2019, "2:07:08", "2:19:10"),
    ("Mainova Frankfurt Marathon", "42K", 2022, "2:06:11", "2:23:11"),
    ("Mainova Frankfurt Marathon", "42K", 2023, "2:04:53", "2:19:27"),
    ("Mainova Frankfurt Marathon", "42K", 2024, "2:05:54", "2:17:25"),
    ("Mainova Frankfurt Marathon", "42K", 2025, "2:06:16", "2:19:34"),
    # Haspa Marathon Hamburg
    ("Haspa Marathon Hamburg", "42K", 2015, "2:07:17", "2:25:41"),
    ("Haspa Marathon Hamburg", "42K", 2016, "2:06:58", "2:21:54"),
    ("Haspa Marathon Hamburg", "42K", 2017, "2:07:26", "2:25:30"),
    ("Haspa Marathon Hamburg", "42K", 2018, "2:06:34", "2:24:51"),
    ("Haspa Marathon Hamburg", "42K", 2019, "2:08:25", "2:24:41"),
    ("Haspa Marathon Hamburg", "42K", 2021, "2:10:14", "2:26:19"),
    ("Haspa Marathon Hamburg", "42K", 2022, "2:04:47", "2:17:23"),
    ("Haspa Marathon Hamburg", "42K", 2023, "2:04:09", "2:20:09"),
    ("Haspa Marathon Hamburg", "42K", 2024, "2:04:24", "2:18:22"),
    ("Haspa Marathon Hamburg", "42K", 2025, "2:03:46", "2:17:55"),
    # Milano Marathon
    ("Milano Marathon", "42K", 2015, "2:08:44", "2:27:35"),
    ("Milano Marathon", "42K", 2016, "2:08:15", "2:27:45"),
    ("Milano Marathon", "42K", 2017, "2:07:13", "2:29:52"),
    ("Milano Marathon", "42K", 2018, "2:09:04", "2:27:02"),
    ("Milano Marathon", "42K", 2019, "2:04:46", "2:22:25"),
    ("Milano Marathon", "42K", 2021, "2:02:57", "2:19:35"),
    ("Milano Marathon", "42K", 2022, "2:05:05", "2:20:17"),
    ("Milano Marathon", "42K", 2023, "2:07:14", "N/A"),
    ("Milano Marathon", "42K", 2024, "2:07:12", "2:26:32"),
    ("Milano Marathon", "42K", 2025, "N/A", "2:23:31"),
    # Brighton Marathon
    ("Brighton Marathon", "42K", 2015, "2:10:14", "2:34:25"),
    ("Brighton Marathon", "42K", 2016, "2:09:56", "2:34:11"),
    ("Brighton Marathon", "42K", 2017, "2:27:36", "2:42:40"),
    ("Brighton Marathon", "42K", 2018, "2:22:34", "2:38:43"),
    ("Brighton Marathon", "42K", 2019, "2:16:23", "2:34:06"),
    ("Brighton Marathon", "42K", 2022, "2:29:07", "2:51:07"),
    ("Brighton Marathon", "42K", 2023, "2:24:07", "2:55:11"),
    ("Brighton Marathon", "42K", 2024, "2:32:27", "2:55:00"),
    ("Brighton Marathon", "42K", 2025, "2:26:47", "2:52:58"),
    # Gold Coast Marathon
    ("Gold Coast Marathon", "42K", 2015, "2:08:42", "N/A"),
    ("Gold Coast Marathon", "42K", 2016, "2:09:00", "2:26:40"),
    ("Gold Coast Marathon", "42K", 2017, "N/A", "2:25:34"),
    ("Gold Coast Marathon", "42K", 2018, "2:09:49", "2:24:49"),
    ("Gold Coast Marathon", "42K", 2019, "2:07:50", "N/A"),
    ("Gold Coast Marathon", "42K", 2022, "N/A", "2:24:43"),
    ("Gold Coast Marathon", "42K", 2023, "2:07:40", "N/A"),
    ("Gold Coast Marathon", "42K", 2024, "N/A", "2:24:22"),
    ("Gold Coast Marathon", "42K", 2025, "2:07:33", "2:29:27"),
    # Portland Marathon
    ("Portland Marathon", "42K", 2015, "2:28:29", "2:51:23"),
    ("Portland Marathon", "42K", 2016, "2:36:25", "2:38:45"),
    ("Portland Marathon", "42K", 2019, "2:25:15", "2:48:00"),
    ("Portland Marathon", "42K", 2021, "2:25:25", "2:39:36"),
    ("Portland Marathon", "42K", 2022, "2:28:53", "2:53:20"),
    ("Portland Marathon", "42K", 2023, "2:27:51", "2:55:50"),
    ("Portland Marathon", "42K", 2024, "2:22:14", "2:48:21"),
    # Columbus Marathon
    ("Columbus Marathon", "42K", 2019, "2:15:05", "2:34:29"),
    ("Columbus Marathon", "42K", 2023, "2:18:07", "N/A"),
    ("Columbus Marathon", "42K", 2025, "2:21:16", "2:42:56"),
    # Richmond Marathon
    ("Richmond Marathon", "42K", 2015, "2:18:48", "2:42:13"),
    ("Richmond Marathon", "42K", 2017, "2:19:44", "2:39:00"),
    ("Richmond Marathon", "42K", 2018, "2:20:44", "2:39:04"),
    ("Richmond Marathon", "42K", 2019, "2:19:43", "2:36:19"),
    ("Richmond Marathon", "42K", 2023, "2:24:18", "2:40:26"),
    # Austin Marathon
    ("Austin Marathon", "42K", 2015, "2:16:20", "2:54:21"),
    ("Austin Marathon", "42K", 2016, "2:23:09", "3:02:38"),
    ("Austin Marathon", "42K", 2017, "2:32:05", "2:48:16"),
    ("Austin Marathon", "42K", 2018, "2:21:37", "2:43:11"),
    ("Austin Marathon", "42K", 2019, "2:17:03", "2:42:27"),
    ("Austin Marathon", "42K", 2022, "2:14:24", "2:35:17"),
    ("Austin Marathon", "42K", 2023, "2:16:35", "2:36:51"),
    ("Austin Marathon", "42K", 2025, "2:18:56", "2:41:04"),
    ("Austin Marathon", "42K", 2026, "2:13:18", "2:33:28"),
    # Detroit Marathon
    ("Detroit Marathon", "42K", 2019, "2:18:58", "2:39:19"),
    ("Detroit Marathon", "42K", 2021, "2:22:03", "2:51:08"),
    ("Detroit Marathon", "42K", 2025, "2:16:10", "2:46:59"),
    # San Francisco Marathon
    ("San Francisco Marathon", "42K", 2015, "2:26:22", "2:49:42"),
    ("San Francisco Marathon", "42K", 2016, "2:30:42", "2:49:51"),
    ("San Francisco Marathon", "42K", 2017, "2:28:23", "2:52:49"),
    ("San Francisco Marathon", "42K", 2018, "2:27:56", "2:54:09"),
    ("San Francisco Marathon", "42K", 2019, "2:25:25", "N/A"),
    ("San Francisco Marathon", "42K", 2021, "2:20:47", "2:55:20"),
    ("San Francisco Marathon", "42K", 2022, "2:31:42", "2:44:38"),
    ("San Francisco Marathon", "42K", 2023, "2:26:17", "2:45:59"),
    ("San Francisco Marathon", "42K", 2024, "2:22:30", "2:52:10"),
    # Flying Pig Marathon
    ("Flying Pig Marathon", "42K", 2015, "2:32:53", "2:53:07"),
    ("Flying Pig Marathon", "42K", 2016, "2:26:03", "2:55:46"),
    ("Flying Pig Marathon", "42K", 2017, "2:33:43", "2:53:47"),
    ("Flying Pig Marathon", "42K", 2018, "2:29:36", "2:46:39"),
    ("Flying Pig Marathon", "42K", 2019, "2:25:14", "2:44:28"),
    ("Flying Pig Marathon", "42K", 2021, "2:28:45", "2:43:44"),
    ("Flying Pig Marathon", "42K", 2022, "2:27:18", "N/A"),
    ("Flying Pig Marathon", "42K", 2023, "2:27:10", "2:45:34"),
    ("Flying Pig Marathon", "42K", 2024, "2:27:36", "2:43:22"),
    ("Flying Pig Marathon", "42K", 2025, "2:22:41", "2:40:06"),
    # Grandma's Marathon
    ("Grandma's Marathon", "42K", 2015, "2:10:38", "2:32:06"),
    ("Grandma's Marathon", "42K", 2016, "2:11:26", "2:33:28"),
    ("Grandma's Marathon", "42K", 2017, "2:12:06", "2:32:09"),
    ("Grandma's Marathon", "42K", 2018, "2:10:06", "2:24:28"),
    ("Grandma's Marathon", "42K", 2019, "2:11:56", "2:28:06"),
    ("Grandma's Marathon", "42K", 2021, "2:13:04", "2:29:04"),
    ("Grandma's Marathon", "42K", 2022, "2:09:34", "2:25:01"),
    ("Grandma's Marathon", "42K", 2023, "2:09:14", "2:25:55"),
    ("Grandma's Marathon", "42K", 2024, "2:10:54", "2:23:52"),
    # Medtronic Twin Cities Marathon
    ("Medtronic Twin Cities Marathon", "42K", 2015, "2:11:16", "2:31:40"),
    ("Medtronic Twin Cities Marathon", "42K", 2016, "2:08:51", "2:30:01"),
    ("Medtronic Twin Cities Marathon", "42K", 2017, "2:11:54", "2:30:26"),
    ("Medtronic Twin Cities Marathon", "42K", 2018, "2:11:58", "2:33:04"),
    ("Medtronic Twin Cities Marathon", "42K", 2019, "2:12:23", "2:31:29"),
    ("Medtronic Twin Cities Marathon", "42K", 2021, "2:15:22", "2:45:55"),
    ("Medtronic Twin Cities Marathon", "42K", 2022, "2:11:28", "2:33:09"),
    ("Medtronic Twin Cities Marathon", "42K", 2024, "2:10:17", "2:28:52"),
    ("Medtronic Twin Cities Marathon", "42K", 2025, "2:15:41", "2:36:43"),
    # St. Jude Memphis Marathon
    ("St. Jude Memphis Marathon", "42K", 2015, "2:38:23", "3:00:20"),
    ("St. Jude Memphis Marathon", "42K", 2023, "2:30:21", "2:56:44"),
    ("St. Jude Memphis Marathon", "42K", 2024, "2:34:18", "2:54:45"),
    ("St. Jude Memphis Marathon", "42K", 2025, "2:28:03", "2:51:12"),
    # California International Marathon
    ("California International Marathon", "42K", 2015, "2:12:12", "2:31:50"),
    ("California International Marathon", "42K", 2016, "2:11:41", "2:31:20"),
    ("California International Marathon", "42K", 2017, "2:11:55", "2:28:10"),
    ("California International Marathon", "42K", 2018, "2:12:39", "2:28:19"),
    ("California International Marathon", "42K", 2019, "2:13:36", "2:29:31"),
    ("California International Marathon", "42K", 2021, "2:11:21", "2:26:53"),
    ("California International Marathon", "42K", 2022, "2:11:01", "2:26:02"),
    ("California International Marathon", "42K", 2023, "2:11:09", "2:29:00"),
    ("California International Marathon", "42K", 2024, "2:07:35", "2:24:28"),
    ("California International Marathon", "42K", 2025, "2:09:29", "2:24:09"),
    # ── AGENT 3: European/other marathons 2000-2014 ──
    # Valencia Marathon 2008-2014
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2008, "2:12:27", "2:36:27"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2009, "2:12:09", "2:35:33"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2010, "2:09:52", "2:33:08"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2011, "2:10:36", "2:35:39"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2012, "2:09:09", "2:31:35"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2013, "2:09:48", "2:32:47"),
    ("Valencia Marathon Trinidad Alfonso Zurich", "42K", 2014, "2:10:42", "2:30:54"),
    # Zurich Maratón de Sevilla 2009-2014
    ("Zurich Maratón de Sevilla", "42K", 2009, "2:11:51", "2:26:03"),
    ("Zurich Maratón de Sevilla", "42K", 2010, "2:09:53", "2:30:55"),
    ("Zurich Maratón de Sevilla", "42K", 2011, "2:10:26", "2:33:10"),
    ("Zurich Maratón de Sevilla", "42K", 2012, "2:10:19", "2:33:41"),
    ("Zurich Maratón de Sevilla", "42K", 2013, "2:11:26", "2:33:44"),
    ("Zurich Maratón de Sevilla", "42K", 2014, "2:08:33", "2:28:39"),
    # Zurich Marató de Barcelona 2006-2014
    ("Zurich Marató de Barcelona", "42K", 2006, "2:11:44", "2:36:04"),
    ("Zurich Marató de Barcelona", "42K", 2007, "2:12:56", "2:37:40"),
    ("Zurich Marató de Barcelona", "42K", 2008, "2:10:55", "2:34:03"),
    ("Zurich Marató de Barcelona", "42K", 2009, "2:10:22", "2:33:19"),
    ("Zurich Marató de Barcelona", "42K", 2010, "2:07:30", "2:31:51"),
    ("Zurich Marató de Barcelona", "42K", 2011, "2:07:47", "2:28:47"),
    ("Zurich Marató de Barcelona", "42K", 2012, "2:10:06", "2:33:10"),
    ("Zurich Marató de Barcelona", "42K", 2013, "2:09:36", "2:31:10"),
    ("Zurich Marató de Barcelona", "42K", 2014, "2:09:41", "2:29:10"),
    # Acea Run Rome The Marathon 2004-2014
    ("Acea Run Rome The Marathon", "42K", 2004, "2:10:12", "2:27:49"),
    ("Acea Run Rome The Marathon", "42K", 2005, "2:08:02", "2:28:01"),
    ("Acea Run Rome The Marathon", "42K", 2006, "2:08:38", "2:25:44"),
    ("Acea Run Rome The Marathon", "42K", 2007, "2:09:36", "2:25:08"),
    ("Acea Run Rome The Marathon", "42K", 2008, "2:09:57", "2:22:53"),
    ("Acea Run Rome The Marathon", "42K", 2009, "2:07:17", "2:27:08"),
    ("Acea Run Rome The Marathon", "42K", 2010, "2:08:39", "2:25:28"),
    ("Acea Run Rome The Marathon", "42K", 2011, "2:08:45", "2:24:13"),
    ("Acea Run Rome The Marathon", "42K", 2012, "2:08:04", "2:31:11"),
    ("Acea Run Rome The Marathon", "42K", 2013, "2:07:56", "2:24:40"),
    ("Acea Run Rome The Marathon", "42K", 2014, "2:09:47", "2:34:49"),
    # Vienna City Marathon 2000-2014
    ("Vienna City Marathon", "42K", 2000, "2:10:24", "2:23:47"),
    ("Vienna City Marathon", "42K", 2001, "2:12:20", "2:31:46"),
    ("Vienna City Marathon", "42K", 2002, "2:10:18", "2:30:41"),
    ("Vienna City Marathon", "42K", 2003, "2:10:28", "2:29:31"),
    ("Vienna City Marathon", "42K", 2004, "2:10:41", "2:29:24"),
    ("Vienna City Marathon", "42K", 2005, "2:12:22", "2:27:33"),
    ("Vienna City Marathon", "42K", 2006, "2:12:09", "2:29:19"),
    ("Vienna City Marathon", "42K", 2007, "2:08:50", "2:30:09"),
    ("Vienna City Marathon", "42K", 2008, "2:07:38", "2:25:47"),
    ("Vienna City Marathon", "42K", 2009, "2:08:21", "2:30:43"),
    ("Vienna City Marathon", "42K", 2010, "2:08:40", "2:31:08"),
    ("Vienna City Marathon", "42K", 2011, "2:08:29", "2:26:21"),
    ("Vienna City Marathon", "42K", 2012, "2:06:58", "2:26:39"),
    ("Vienna City Marathon", "42K", 2013, "2:08:40", "2:31:08"),
    ("Vienna City Marathon", "42K", 2014, "2:05:41", "2:28:59"),
    # Copenhagen Marathon 2000-2014
    ("Copenhagen Marathon", "42K", 2000, "2:21:30", "2:47:42"),
    ("Copenhagen Marathon", "42K", 2001, "2:22:15", "2:47:08"),
    ("Copenhagen Marathon", "42K", 2002, "2:23:53", "2:46:30"),
    ("Copenhagen Marathon", "42K", 2003, "2:19:57", "2:44:30"),
    ("Copenhagen Marathon", "42K", 2004, "2:23:10", "2:49:55"),
    ("Copenhagen Marathon", "42K", 2005, "2:21:14", "2:44:30"),
    ("Copenhagen Marathon", "42K", 2006, "2:24:08", "2:48:22"),
    ("Copenhagen Marathon", "42K", 2007, "2:23:54", "2:40:21"),
    ("Copenhagen Marathon", "42K", 2008, "2:21:05", "2:41:14"),
    ("Copenhagen Marathon", "42K", 2009, "2:18:04", "2:41:00"),
    ("Copenhagen Marathon", "42K", 2010, "2:22:29", "2:38:49"),
    ("Copenhagen Marathon", "42K", 2011, "2:21:45", "2:45:30"),
    ("Copenhagen Marathon", "42K", 2012, "2:24:49", "2:41:17"),
    ("Copenhagen Marathon", "42K", 2013, "2:17:24", "2:44:12"),
    ("Copenhagen Marathon", "42K", 2014, "2:21:38", "2:40:02"),
    # Prague International Marathon 2000-2014
    ("Prague International Marathon", "42K", 2000, "2:10:35", "2:30:29"),
    ("Prague International Marathon", "42K", 2001, "2:12:48", "2:30:35"),
    ("Prague International Marathon", "42K", 2002, "2:11:41", "2:32:24"),
    ("Prague International Marathon", "42K", 2003, "2:11:56", "2:31:10"),
    ("Prague International Marathon", "42K", 2004, "2:12:15", "2:31:48"),
    ("Prague International Marathon", "42K", 2005, "2:10:51", "2:29:42"),
    ("Prague International Marathon", "42K", 2006, "2:11:11", "2:32:14"),
    ("Prague International Marathon", "42K", 2007, "2:11:49", "2:30:11"),
    ("Prague International Marathon", "42K", 2008, "2:11:06", "2:30:09"),
    ("Prague International Marathon", "42K", 2009, "2:10:42", "2:32:33"),
    ("Prague International Marathon", "42K", 2010, "2:05:39", "2:25:19"),
    ("Prague International Marathon", "42K", 2011, "2:07:07", "2:22:34"),
    ("Prague International Marathon", "42K", 2012, "2:08:11", "2:27:21"),
    ("Prague International Marathon", "42K", 2013, "2:09:30", "2:27:32"),
    ("Prague International Marathon", "42K", 2014, "2:10:18", "2:27:15"),
    # Irish Life Dublin Marathon 2000-2014
    ("Irish Life Dublin Marathon", "42K", 2000, "2:18:49", "2:35:42"),
    ("Irish Life Dublin Marathon", "42K", 2001, "2:13:59", "2:36:12"),
    ("Irish Life Dublin Marathon", "42K", 2002, "2:14:25", "2:32:58"),
    ("Irish Life Dublin Marathon", "42K", 2003, "2:14:25", "2:27:22"),
    ("Irish Life Dublin Marathon", "42K", 2004, "2:13:08", "2:32:53"),
    ("Irish Life Dublin Marathon", "42K", 2005, "2:13:14", "2:34:45"),
    ("Irish Life Dublin Marathon", "42K", 2006, "2:11:39", "2:29:49"),
    ("Irish Life Dublin Marathon", "42K", 2007, "2:09:07", "2:30:32"),
    ("Irish Life Dublin Marathon", "42K", 2008, "2:11:04", "2:33:49"),
    ("Irish Life Dublin Marathon", "42K", 2009, "2:09:12", "2:32:45"),
    ("Irish Life Dublin Marathon", "42K", 2010, "2:08:58", "2:26:13"),
    ("Irish Life Dublin Marathon", "42K", 2011, "2:08:33", "2:28:51"),
    ("Irish Life Dublin Marathon", "42K", 2012, "2:13:06", "2:31:22"),
    ("Irish Life Dublin Marathon", "42K", 2013, "2:11:09", "2:30:46"),
    ("Irish Life Dublin Marathon", "42K", 2014, "2:14:48", "2:34:15"),
    # Marine Corps Marathon 2000-2014
    ("Marine Corps Marathon", "42K", 2000, "2:22:13", "2:47:46"),
    ("Marine Corps Marathon", "42K", 2001, "2:28:28", "2:48:13"),
    ("Marine Corps Marathon", "42K", 2002, "2:23:02", "2:48:40"),
    ("Marine Corps Marathon", "42K", 2003, "2:22:11", "2:37:59"),
    ("Marine Corps Marathon", "42K", 2004, "2:23:49", "2:48:31"),
    ("Marine Corps Marathon", "42K", 2005, "2:22:18", "2:49:04"),
    ("Marine Corps Marathon", "42K", 2006, "2:21:20", "2:45:17"),
    ("Marine Corps Marathon", "42K", 2007, "2:20:17", "2:41:58"),
    ("Marine Corps Marathon", "42K", 2008, "2:21:58", "2:45:21"),
    ("Marine Corps Marathon", "42K", 2009, "2:24:43", "2:48:17"),
    ("Marine Corps Marathon", "42K", 2010, "2:21:32", "2:39:19"),
    ("Marine Corps Marathon", "42K", 2011, "2:21:30", "2:38:57"),
    ("Marine Corps Marathon", "42K", 2013, "2:23:37", "2:42:09"),
    ("Marine Corps Marathon", "42K", 2014, "2:23:53", "2:51:46"),
    # ASICS Los Angeles Marathon 2009-2014
    ("ASICS Los Angeles Marathon", "42K", 2009, "2:08:24", "2:25:59"),
    ("ASICS Los Angeles Marathon", "42K", 2010, "2:09:19", "2:25:38"),
    ("ASICS Los Angeles Marathon", "42K", 2011, "2:06:35", "2:26:34"),
    ("ASICS Los Angeles Marathon", "42K", 2012, "2:12:19", "2:25:39"),
    ("ASICS Los Angeles Marathon", "42K", 2013, "2:09:44", "2:26:05"),
    ("ASICS Los Angeles Marathon", "42K", 2014, "2:10:37", "2:27:37"),
    # Chevron Houston Marathon 2000-2014
    ("Chevron Houston Marathon", "42K", 2000, "2:11:27", "2:32:24"),
    ("Chevron Houston Marathon", "42K", 2001, "2:29:27", "2:43:41"),
    ("Chevron Houston Marathon", "42K", 2002, "2:28:43", "2:50:49"),
    ("Chevron Houston Marathon", "42K", 2003, "2:24:43", "2:42:37"),
    ("Chevron Houston Marathon", "42K", 2004, "2:18:51", "2:28:36"),
    ("Chevron Houston Marathon", "42K", 2005, "2:14:50", "2:32:27"),
    ("Chevron Houston Marathon", "42K", 2006, "2:12:02", "2:32:25"),
    ("Chevron Houston Marathon", "42K", 2007, "2:11:39", "2:26:52"),
    ("Chevron Houston Marathon", "42K", 2008, "2:12:32", "2:24:40"),
    ("Chevron Houston Marathon", "42K", 2009, "2:07:52", "2:24:18"),
    ("Chevron Houston Marathon", "42K", 2010, "2:07:37", "2:23:53"),
    ("Chevron Houston Marathon", "42K", 2011, "2:07:04", "2:26:33"),
    ("Chevron Houston Marathon", "42K", 2012, "2:06:51", "2:23:14"),
    ("Chevron Houston Marathon", "42K", 2013, "2:10:17", "2:23:37"),
    ("Chevron Houston Marathon", "42K", 2014, "2:07:32", "2:25:52"),
]

# ============================================================
# SEMI-MARATHONS (21K)
# ============================================================

half_marathon_data = [
    # HOKA Semi de Paris
    ("HOKA Semi de Paris", "21K", 2023, "0:59:38", "1:06:01"),
    ("HOKA Semi de Paris", "21K", 2024, "1:00:45", "1:06:58"),
    ("HOKA Semi de Paris", "21K", 2025, "1:00:16", "1:07:14"),
    ("HOKA Semi de Paris", "21K", 2026, "1:00:11", "1:05:12"),
    # Mitja Marato Barcelona by Brooks
    ("Mitja Marato Barcelona by Brooks", "21K", 2023, "0:58:53", "1:04:37"),
    ("Mitja Marato Barcelona by Brooks", "21K", 2024, "0:59:21", "1:04:28"),
    ("Mitja Marato Barcelona by Brooks", "21K", 2025, "0:56:42", "1:04:11"),
    ("Mitja Marato Barcelona by Brooks", "21K", 2026, "N/A", "1:04:00"),
    # Aramco Houston Half Marathon
    ("Aramco Houston Half Marathon", "21K", 2024, "1:00:42", "1:04:37"),
    ("Aramco Houston Half Marathon", "21K", 2025, "0:59:17", "N/A"),
    ("Aramco Houston Half Marathon", "21K", 2026, "0:59:01", "1:04:49"),
    # Medio Maraton de Sevilla
    ("Medio Maraton de Sevilla", "21K", 2024, "0:59:21", "1:07:59"),
    ("Medio Maraton de Sevilla", "21K", 2025, "0:59:33", "1:07:18"),
    ("Medio Maraton de Sevilla", "21K", 2026, "1:00:24", "1:06:33"),
    # Standard Chartered Hong Kong Half Marathon
    ("Standard Chartered Hong Kong Half Marathon", "21K", 2025, "N/A", "N/A"),
    # Walt Disney World Marathon Weekend (Half)
    ("Walt Disney World Marathon Weekend", "21K", 2025, "N/A", "N/A"),
    ("Walt Disney World Marathon Weekend", "21K", 2026, "N/A", "N/A"),
    # Tata Mumbai Marathon (Half)
    ("Tata Mumbai Marathon (Half)", "21K", 2025, "N/A", "N/A"),
    ("Tata Mumbai Marathon (Half)", "21K", 2026, "N/A", "N/A"),
    # Riyadh Marathon (Half)
    ("Riyadh Marathon (Half)", "21K", 2023, "N/A", "N/A"),
    ("Riyadh Marathon (Half)", "21K", 2024, "N/A", "N/A"),
    ("Riyadh Marathon (Half)", "21K", 2025, "N/A", "N/A"),
    ("Riyadh Marathon (Half)", "21K", 2026, "N/A", "N/A"),
    # Rock 'n' Roll Running Series Las Vegas
    ("Rock 'n' Roll Running Series Las Vegas", "21K", 2023, "N/A", "N/A"),
    ("Rock 'n' Roll Running Series Las Vegas", "21K", 2024, "N/A", "N/A"),
    ("Rock 'n' Roll Running Series Las Vegas", "21K", 2025, "N/A", "N/A"),
    ("Rock 'n' Roll Running Series Las Vegas", "21K", 2026, "N/A", "N/A"),
    # Burj2Burj Half Marathon
    ("Burj2Burj Half Marathon", "21K", 2024, "N/A", "N/A"),
    ("Burj2Burj Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Burj2Burj Half Marathon", "21K", 2026, "0:59:26", "1:06:57"),
    # Generali Berlin Half Marathon
    ("Generali Berlin Half Marathon", "21K", 2023, "0:59:01", "1:05:43"),
    ("Generali Berlin Half Marathon", "21K", 2024, "0:59:30", "1:06:53"),
    ("Generali Berlin Half Marathon", "21K", 2025, "0:58:43", "1:03:35"),
    # United Airlines NYC Half
    ("United Airlines NYC Half", "21K", 2025, "1:01:31", "1:07:04"),
    ("United Airlines NYC Half", "21K", 2026, "0:59:30", "1:06:33"),
    # EDP Lisboa Meia Maratona
    ("EDP Lisboa Meia Maratona", "21K", 2024, "N/A", "N/A"),
    ("EDP Lisboa Meia Maratona", "21K", 2025, "N/A", "N/A"),
    ("EDP Lisboa Meia Maratona", "21K", 2026, "N/A", "N/A"),
    # Generali Prague Half Marathon
    ("Generali Prague Half Marathon", "21K", 2023, "0:59:43", "1:06:00"),
    ("Generali Prague Half Marathon", "21K", 2024, "0:58:24", "1:08:10"),
    ("Generali Prague Half Marathon", "21K", 2025, "0:58:54", "1:05:27"),
    # Eurospin RomaOstia Half Marathon
    ("Eurospin RomaOstia Half Marathon", "21K", 2023, "0:59:17", "1:06:21"),
    ("Eurospin RomaOstia Half Marathon", "21K", 2024, "1:01:10", "1:07:38"),
    ("Eurospin RomaOstia Half Marathon", "21K", 2025, "0:58:49", "1:08:20"),
    ("Eurospin RomaOstia Half Marathon", "21K", 2026, "0:58:00", "N/A"),
    # Copenhagen Half Marathon
    ("Copenhagen Half Marathon", "21K", 2023, "0:59:11", "1:05:53"),
    ("Copenhagen Half Marathon", "21K", 2024, "0:58:05", "1:05:11"),
    ("Copenhagen Half Marathon", "21K", 2025, "0:58:23", "1:04:44"),
    # AJ Bell Great North Run
    ("AJ Bell Great North Run", "21K", 2025, "1:00:52", "1:09:32"),
    # Valencia Half Marathon Trinidad Alfonso Zurich
    ("Valencia Half Marathon Trinidad Alfonso Zurich", "21K", 2025, "0:58:02", "1:03:08"),
    # Göteborgsvarvet
    ("Göteborgsvarvet", "21K", 2024, "N/A", "N/A"),
    ("Göteborgsvarvet", "21K", 2025, "N/A", "N/A"),
    # Run in Lyon (Semi)
    ("Run in Lyon (Semi)", "21K", 2023, "N/A", "N/A"),
    ("Run in Lyon (Semi)", "21K", 2024, "1:05:44", "1:18:34"),
    ("Run in Lyon (Semi)", "21K", 2025, "N/A", "N/A"),
    # Zurich Rock 'n' Roll Running Series Madrid (Half)
    ("Zurich Rock 'n' Roll Running Series Madrid (Half)", "21K", 2023, "N/A", "N/A"),
    ("Zurich Rock 'n' Roll Running Series Madrid (Half)", "21K", 2024, "N/A", "N/A"),
    ("Zurich Rock 'n' Roll Running Series Madrid (Half)", "21K", 2025, "N/A", "N/A"),
    # Warsaw Half Marathon
    ("Warsaw Half Marathon", "21K", 2024, "N/A", "N/A"),
    ("Warsaw Half Marathon", "21K", 2025, "N/A", "N/A"),
    # TCS Toronto Waterfront Marathon (Half)
    ("TCS Toronto Waterfront Marathon (Half)", "21K", 2023, "N/A", "N/A"),
    ("TCS Toronto Waterfront Marathon (Half)", "21K", 2024, "N/A", "N/A"),
    ("TCS Toronto Waterfront Marathon (Half)", "21K", 2025, "N/A", "N/A"),
    # Hyundai Meia Maratona (Lisbon)
    ("Hyundai Meia Maratona", "21K", 2023, "N/A", "N/A"),
    ("Hyundai Meia Maratona", "21K", 2024, "N/A", "N/A"),
    ("Hyundai Meia Maratona", "21K", 2025, "N/A", "N/A"),
    # Taipei Half Marathon
    ("Taipei Half Marathon", "21K", 2023, "N/A", "N/A"),
    ("Taipei Half Marathon", "21K", 2024, "N/A", "N/A"),
    ("Taipei Half Marathon", "21K", 2025, "N/A", "N/A"),
    # Bangsaen21
    ("Bangsaen21", "21K", 2023, "N/A", "N/A"),
    ("Bangsaen21", "21K", 2024, "N/A", "N/A"),
    ("Bangsaen21", "21K", 2025, "N/A", "N/A"),
    # Other half marathons with limited data (1 year only, 2025)
    ("Coelmo Napoli City Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Coelmo Napoli City Half Marathon", "21K", 2026, "N/A", "N/A"),
    ("Kagawa Marugame International Half Marathon", "21K", 2026, "N/A", "N/A"),
    ("Movistar Madrid Medio Maraton", "21K", 2025, "N/A", "N/A"),
    ("NN CPC Loop Den Haag - Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("NN CPC Loop Den Haag - Half Marathon", "21K", 2026, "N/A", "N/A"),
    ("Disney Princess Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Disney Princess Half Marathon", "21K", 2026, "N/A", "N/A"),
    ("St. Jude Rock 'n' Roll Running Series Washington DC", "21K", 2025, "N/A", "N/A"),
    ("GetPro Bath Half", "21K", 2025, "N/A", "N/A"),
    ("GetPro Bath Half", "21K", 2026, "N/A", "N/A"),
    ("NYCRUNS Brooklyn Experience Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("London Landmarks Half Marathon", "21K", 2024, "N/A", "N/A"),
    ("London Landmarks Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Vienna City Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("21K de Montréal", "21K", 2025, "N/A", "N/A"),
    ("St. Jude Rock 'n' Roll Series Nashville", "21K", 2025, "N/A", "N/A"),
    ("NYRR RBC Brooklyn Half", "21K", 2025, "N/A", "N/A"),
    ("Hoka Runaway Sydney Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("IU Health 500 Festival Mini-Marathon", "21K", 2025, "N/A", "N/A"),
    ("Rock 'n' Roll Running Series San Diego", "21K", 2025, "N/A", "N/A"),
    ("AJ Bell Great Manchester Run", "21K", 2025, "N/A", "N/A"),
    ("AJ Bell Great Bristol Run", "21K", 2025, "N/A", "N/A"),
    ("Brølløbet - The Bridge Run", "21K", 2025, "N/A", "N/A"),
    ("Media Maraton de Bogota", "21K", 2025, "N/A", "N/A"),
    ("Asics Run Melbourne", "21K", 2025, "N/A", "N/A"),
    ("The Big Half", "21K", 2025, "N/A", "N/A"),
    ("Life Time Chicago Half Marathon & 5K", "21K", 2024, "N/A", "N/A"),
    ("Life Time Chicago Half Marathon & 5K", "21K", 2025, "N/A", "N/A"),
    ("Cardiff Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Manchester Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Wizz Air Rome Half Marathon by Brooks", "21K", 2025, "N/A", "N/A"),
    ("Royal Parks Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Tokyo Legacy Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("TCS Mizuno Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Nike Melbourne Marathon Festival", "21K", 2025, "N/A", "N/A"),
    ("Vedanta Delhi Half Marathon", "21K", 2025, "0:59:50", "1:07:20"),
    ("Half Marathon Munchen by Brooks", "21K", 2025, "N/A", "N/A"),
    ("Semi-Marathon de Boulogne-Billancourt", "21K", 2025, "N/A", "N/A"),
    ("Standard Chartered Singapore Half Marathon", "21K", 2025, "N/A", "N/A"),
    ("Rock 'n' Roll Running Series Mexico City", "21K", 2025, "N/A", "N/A"),
    # Standard Chartered KL Half Marathon
    ("Standard Chartered KL Half Marathon", "21K", 2025, "1:11:08", "1:23:55"),
    # Standard Chartered Hong Kong Half Marathon
    ("Standard Chartered Hong Kong Half Marathon", "21K", 2026, "N/A", "N/A"),
    ("Standard Chartered Hong Kong Half Marathon", "21K", 2019, "1:01:13", "1:09:28"),
]

# ============================================================
# 10K
# ============================================================

tenk_data = [
    # 10K Valencia Ibercaja by Kiprun
    ("10K Valencia Ibercaja by Kiprun", "10K", 2023, "N/A", "0:29:19"),
    ("10K Valencia Ibercaja by Kiprun", "10K", 2024, "N/A", "0:28:45"),
    ("10K Valencia Ibercaja by Kiprun", "10K", 2025, "N/A", "N/A"),
    ("10K Valencia Ibercaja by Kiprun", "10K", 2026, "0:26:45", "0:29:25"),
    # 10K Montmartre
    ("10K Montmartre", "10K", 2026, "N/A", "N/A"),
    # Cancer Research UK London Winter Run
    ("Cancer Research UK London Winter Run", "10K", 2023, "0:28:52", "0:31:59"),
    ("Cancer Research UK London Winter Run", "10K", 2024, "0:28:46", "0:32:33"),
    ("Cancer Research UK London Winter Run", "10K", 2025, "0:29:22", "0:34:32"),
    ("Cancer Research UK London Winter Run", "10K", 2026, "N/A", "N/A"),
    # EDP Meia Maratona de Lisboa - Vodafone 10K
    ("EDP Meia Maratona de Lisboa - Vodafone 10K", "10K", 2023, "N/A", "N/A"),
    ("EDP Meia Maratona de Lisboa - Vodafone 10K", "10K", 2024, "N/A", "N/A"),
    ("EDP Meia Maratona de Lisboa - Vodafone 10K", "10K", 2025, "N/A", "N/A"),
    ("EDP Meia Maratona de Lisboa - Vodafone 10K", "10K", 2026, "N/A", "N/A"),
    # Cooper River Bridge Run
    ("Cooper River Bridge Run", "10K", 2023, "N/A", "N/A"),
    ("Cooper River Bridge Run", "10K", 2024, "N/A", "N/A"),
    ("Cooper River Bridge Run", "10K", 2025, "0:28:27", "0:31:49"),
    # Vancouver Sun Run
    ("Vancouver Sun Run", "10K", 2023, "N/A", "N/A"),
    ("Vancouver Sun Run", "10K", 2024, "N/A", "N/A"),
    ("Vancouver Sun Run", "10K", 2025, "0:28:09", "0:32:54"),
    # Statesman Capitol 10K
    ("Statesman Capitol 10K", "10K", 2023, "N/A", "N/A"),
    ("Statesman Capitol 10K", "10K", 2024, "N/A", "N/A"),
    ("Statesman Capitol 10K", "10K", 2025, "N/A", "N/A"),
    # Ukrop's Monument Avenue 10K
    ("Ukrop's Monument Avenue 10K", "10K", 2023, "N/A", "N/A"),
    ("Ukrop's Monument Avenue 10K", "10K", 2024, "N/A", "N/A"),
    ("Ukrop's Monument Avenue 10K", "10K", 2025, "N/A", "N/A"),
    # Bangsaen10
    ("Bangsaen10", "10K", 2023, "N/A", "N/A"),
    ("Bangsaen10", "10K", 2024, "N/A", "N/A"),
    ("Bangsaen10", "10K", 2025, "N/A", "N/A"),
    # AJ Bell Great Manchester Run (10K)
    ("AJ Bell Great Manchester Run (10K)", "10K", 2024, "N/A", "N/A"),
    ("AJ Bell Great Manchester Run (10K)", "10K", 2025, "N/A", "N/A"),
    # BOLDERBoulder 10K
    ("BOLDERBoulder 10K", "10K", 2023, "0:29:08", "0:33:25"),
    ("BOLDERBoulder 10K", "10K", 2024, "0:29:13", "0:32:45"),
    ("BOLDERBoulder 10K", "10K", 2025, "N/A", "N/A"),
    # Adidas 10K Paris
    ("Adidas 10K Paris", "10K", 2023, "N/A", "N/A"),
    ("Adidas 10K Paris", "10K", 2024, "0:29:37", "0:32:18"),
    ("Adidas 10K Paris", "10K", 2025, "N/A", "N/A"),
    # Royal Run
    ("Royal Run", "10K", 2023, "N/A", "N/A"),
    ("Royal Run", "10K", 2024, "N/A", "N/A"),
    ("Royal Run", "10K", 2025, "N/A", "N/A"),
    # Atlanta Journal-Constitution Peachtree Road Race
    ("Atlanta Journal-Constitution Peachtree Road Race", "10K", 2023, "0:27:41", "0:30:43"),
    ("Atlanta Journal-Constitution Peachtree Road Race", "10K", 2024, "0:28:03", "0:31:12"),
    ("Atlanta Journal-Constitution Peachtree Road Race", "10K", 2025, "N/A", "0:31:29"),
    # Saucony London 10K
    ("Saucony London 10K", "10K", 2023, "N/A", "N/A"),
    ("Saucony London 10K", "10K", 2024, "0:29:41", "0:32:51"),
    ("Saucony London 10K", "10K", 2025, "0:29:33", "N/A"),
    # Transurban Bridge to Brisbane
    ("Transurban Bridge to Brisbane", "10K", 2023, "N/A", "N/A"),
    ("Transurban Bridge to Brisbane", "10K", 2024, "N/A", "N/A"),
    ("Transurban Bridge to Brisbane", "10K", 2025, "N/A", "N/A"),
    # Vitality London 10,000
    ("Vitality London 10,000", "10K", 2023, "N/A", "N/A"),
    ("Vitality London 10,000", "10K", 2024, "0:29:14", "0:31:36"),
    ("Vitality London 10,000", "10K", 2025, "N/A", "N/A"),
    # Run in Lyon (10K)
    ("Run in Lyon (10K)", "10K", 2023, "N/A", "N/A"),
    ("Run in Lyon (10K)", "10K", 2024, "N/A", "N/A"),
    ("Run in Lyon (10K)", "10K", 2025, "N/A", "N/A"),
    # ASICS LDNX
    ("ASICS LDNX", "10K", 2025, "N/A", "N/A"),
    # NN San Silvestre Vallecana
    ("NN San Silvestre Vallecana", "10K", 2023, "N/A", "N/A"),  # Aregawi / Garcia - temps exacts non trouvés
    ("NN San Silvestre Vallecana", "10K", 2024, "N/A", "0:31:18"),
    ("NN San Silvestre Vallecana", "10K", 2025, "0:27:41", "0:31:11"),
    # NN CPC Loop Den Haag - 10 KM Loop
    ("NN CPC Loop Den Haag - 10 KM Loop", "10K", 2026, "N/A", "N/A"),
    # BOLDERBoulder 10K
    ("BOLDERBoulder 10K", "10K", 2025, "0:28:36", "0:31:51"),
    # Cooper River Bridge Run
    ("Cooper River Bridge Run", "10K", 2025, "N/A", "N/A"),
    # Standard Chartered KL 10K
    ("Standard Chartered KL 10K", "10K", 2025, "0:32:26", "0:39:23"),
    # Great Scottish Run 10K
    ("Great Scottish Run 10K", "10K", 2025, "N/A", "N/A"),
    # AJ Bell Great Bristol Run 10K
    ("AJ Bell Great Bristol Run 10K", "10K", 2025, "N/A", "N/A"),
    # AJ Bell Great Birmingham Run 10K
    ("AJ Bell Great Birmingham Run 10K", "10K", 2025, "N/A", "N/A"),
    # AJ Bell Great North 10K
    ("AJ Bell Great North 10K", "10K", 2025, "N/A", "N/A"),
    # AJ Bell Great Manchester Run 10K
    ("AJ Bell Great Manchester Run", "10K", 2025, "N/A", "N/A"),
    # Run in Lyon 10K
    ("Run in Lyon", "10K", 2025, "N/A", "N/A"),
]

# ============================================================
# AUTRES (non-standard distances)
# ============================================================

autre_data = [
    # Dam tot Damloop (10 miles / 16.1km)
    ("Dam tot Damloop", "AUTRE", 2019, "0:45:15", "0:50:45"),
    ("Dam tot Damloop", "AUTRE", 2024, "0:44:51", "0:51:15"),
    ("Dam tot Damloop", "AUTRE", 2025, "0:46:07", "0:50:51"),
    # Bay to Breakers (12K)
    ("Bay to Breakers", "AUTRE", 2019, "0:35:01", "0:39:28"),
    ("Bay to Breakers", "AUTRE", 2022, "0:36:10", "0:42:05"),
    ("Bay to Breakers", "AUTRE", 2024, "0:37:02", "0:43:48"),
    # Broad Street Run (10 miles)
    ("Broad Street Run", "AUTRE", 2025, "0:46:13", "0:54:01"),
    # Boilermaker Road Race (15K)
    ("Boilermaker Road Race", "AUTRE", 2025, "0:42:44", "N/A"),
    # AJ Bell Great South Run (10 miles)
    ("AJ Bell Great South Run", "AUTRE", 2025, "N/A", "N/A"),
    # Falmouth Road Race (7 miles)
    ("Falmouth Road Race", "AUTRE", 2025, "N/A", "N/A"),
    # City2Surf (14K)
    ("City2Surf", "AUTRE", 2024, "N/A", "N/A"),
    # Lilac Bloomsday Run (12K)
    ("Lilac Bloomsday Run", "AUTRE", 2025, "N/A", "N/A"),
    # Manchester Road Race (4.748 miles)
    ("Manchester Road Race", "AUTRE", 2024, "N/A", "N/A"),
    # Army Ten Miler
    ("Army Ten Miler", "AUTRE", 2025, "N/A", "N/A"),
    # Great Scottish Run HM
    ("Great Scottish Run", "SEMI", 2025, "N/A", "N/A"),
    # AJ Bell Great Birmingham Run HM
    ("AJ Bell Great Birmingham Run", "SEMI", 2025, "N/A", "N/A"),
]

data.extend(autre_data)

data.extend(marathon_data)
data.extend(half_marathon_data)
data.extend(tenk_data)

# Sort by Distance (42K first, then 21K, then 10K), then Course, then Year
distance_order = {"42K": 0, "21K": 1, "10K": 2}
data.sort(key=lambda x: (distance_order.get(x[1], 9), x[0], x[2]))

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Chronos Vainqueurs"

# Header style
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Distance color fills
fill_42k = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # Blue
fill_21k = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Green
fill_10k = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Orange
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Headers
headers = ["Course", "Distance", "Année", "Temps Vainqueur Homme", "Temps Vainqueur Femme"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Data rows
for row_idx, (course, distance, year, men_time, women_time) in enumerate(data, 2):
    ws.cell(row=row_idx, column=1, value=course).border = thin_border
    ws.cell(row=row_idx, column=2, value=distance).border = thin_border
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=3, value=year).border = thin_border
    ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=4, value=men_time).border = thin_border
    ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=5, value=women_time).border = thin_border
    ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")

    # Color by distance
    if distance == "42K":
        fill = fill_42k if row_idx % 2 == 0 else fill_white
    elif distance == "21K":
        fill = fill_21k if row_idx % 2 == 0 else fill_white
    else:
        fill = fill_10k if row_idx % 2 == 0 else fill_white

    for col in range(1, 6):
        ws.cell(row=row_idx, column=col).fill = fill

# Column widths
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 28

# Freeze panes
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:E{len(data) + 1}"

# Save
output_path = r"C:\Users\mathi\Documents\6. DATA PACE\0. Dashboard\Fichiers sources\Chronos_Vainqueurs.xlsx"
wb.save(output_path)

# Stats
total = len(data)
with_data = sum(1 for r in data if r[3] != "N/A" or r[4] != "N/A")
marathons = sum(1 for r in data if r[1] == "42K")
halfs = sum(1 for r in data if r[1] == "21K")
tenks = sum(1 for r in data if r[1] == "10K")
courses = len(set(r[0] for r in data))

print(f"Fichier sauvegardé: {output_path}")
print(f"Total lignes: {total}")
print(f"  - 42K: {marathons} lignes")
print(f"  - 21K: {halfs} lignes")
print(f"  - 10K: {tenks} lignes")
print(f"Courses uniques: {courses}")
print(f"Lignes avec au moins 1 chrono: {with_data}/{total} ({100*with_data//total}%)")
